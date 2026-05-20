"""
player_profile_store.py
=======================
Manages named player profiles for the Clone AI system.

Each player gets a JSON file on disk that stores every round they have
ever played, along with derived statistics (gesture frequencies, move
transitions, response tendencies after win/loss/draw). Those stats are
then used by the clone AI to reproduce the player's play-style.

A combined Excel workbook is also kept for research purposes — every
round from every player lands in the "All_Rounds" sheet, and each player
with enough data gets their own analysis tab.

Why JSON + Excel?
    JSON is the authoritative source of truth (fast reads, safe to update).
    Excel is supplementary — useful for analysis and presentations, but if
    the file breaks we just regenerate it from the JSON files.

Storage layout:
    ~/Desktop/CapStone/player_profiles/<name>.json   -- per-player history
    ~/Desktop/CapStone/player_research_log.xlsx      -- combined research log
"""

import csv
import json
from datetime import datetime
from pathlib import Path

# Try to import openpyxl for Excel support. If it isn't installed, all Excel
# writes are silently skipped rather than crashing the game.
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# The three gestures the game recognises.
GESTURES = ["Rock", "Paper", "Scissors"]

# The three possible round outcomes from the player's perspective.
OUTCOMES = ["win", "lose", "draw"]

# UPGRADE: the gesture that "beats" X in the cycle. e.g. Rock -> Paper.
UPGRADE = {"Rock": "Paper", "Paper": "Scissors", "Scissors": "Rock"}

# DOWNGRADE: the gesture that X beats. e.g. Rock -> Scissors.
DOWNGRADE = {"Rock": "Scissors", "Paper": "Rock", "Scissors": "Paper"}

# How many rounds a player needs before the clone AI can reliably mimic them.
MIN_ROUNDS_FOR_CLONE = 30


class PlayerProfileStore:
    """
    Handles all read/write operations for player profile data.

    One instance of this lives in the game and is shared across modes.
    Any mode that wants to record rounds or load clone data goes through here.
    """

    def __init__(self, base_dir=None):
        # Default storage location: ~/Desktop/CapStone/player_profiles/
        self.base_dir     = Path(base_dir) if base_dir else Path.home() / "Desktop" / "CapStone"
        self.profiles_dir = self.base_dir / "player_profiles"
        # Create the full directory tree if it doesn't already exist.
        self.profiles_dir.mkdir(parents=True, exist_ok=True)
        self.excel_path = self.base_dir / "player_research_log.xlsx"

    # ------------------------------------------------------------------
    # File path helpers
    # ------------------------------------------------------------------

    def _profile_path(self, name):
        """
        Convert a player display name to a safe lowercase JSON filename.
        Replaces any filesystem-unfriendly characters with underscores.
        e.g. "Alice B." -> "alice_b_.json"
        """
        safe_name = "".join(c if c.isalnum() or c in "._-" else "_" for c in name)
        return self.profiles_dir / f"{safe_name.lower()}.json"

    # ------------------------------------------------------------------
    # Listing players
    # ------------------------------------------------------------------

    def list_players(self):
        """
        Scan the profiles directory and return basic info about every saved player.

        Returns a list of (display_name, round_count) tuples, sorted alphabetically
        by filename. Silently skips files that are missing, corrupted, or unreadable.
        """
        players = []
        for path in sorted(self.profiles_dir.glob("*.json")):
            try:
                with open(path, "r") as f:
                    data = json.load(f)
                name  = data.get("player_name", path.stem)
                count = len(data.get("rounds", []))
                players.append((name, count))
            except Exception:
                # Bad JSON — skip it silently, don't crash the whole list.
                continue
        return players

    def list_playable_clones(self):
        """
        Return just the names of players who have enough data to clone.
        We need at least MIN_ROUNDS_FOR_CLONE rounds before the stats are meaningful.
        """
        return [
            name for name, count in self.list_players()
            if count >= MIN_ROUNDS_FOR_CLONE
        ]

    # ------------------------------------------------------------------
    # Loading and creating profiles
    # ------------------------------------------------------------------

    def load_profile(self, name):
        """
        Load a player's JSON profile from disk.
        Returns the profile dict, or None if the file doesn't exist or is corrupted.
        """
        path = self._profile_path(name)
        if not path.exists():
            return None
        try:
            with open(path, "r") as f:
                return json.load(f)
        except Exception:
            return None

    def get_or_create_profile(self, name):
        """
        Load the player's existing profile, or create a fresh empty one if they're new.
        Safe to call every round — it won't overwrite existing data.
        """
        profile = self.load_profile(name)
        if profile is not None:
            return profile

        # First time this player has been seen — build a blank profile.
        profile = {
            "player_name": name,
            "created_at":  datetime.now().isoformat(),
            "last_played": datetime.now().isoformat(),
            "rounds":      [],
        }
        self._save_profile(name, profile)
        return profile

    # ------------------------------------------------------------------
    # Recording rounds
    # ------------------------------------------------------------------

    def record_round(self, player_name, player_gesture, robot_gesture,
                     outcome, game_mode, round_number=0, emotion=None,
                     reaction_ms=None):
        """
        Append one round of gameplay to the player's profile and the Excel log.

        Called automatically during gameplay whenever a player name is set.
        Anonymous players (empty name) are silently skipped.

        Parameters
        ----------
        player_name    : str  -- the named player (empty = don't record)
        player_gesture : str  -- what the player threw ("Rock" / "Paper" / "Scissors")
        robot_gesture  : str  -- what the AI threw
        outcome        : str  -- "win" / "lose" / "draw" from the player's view
        game_mode      : str  -- which game mode was active, e.g. "FairPlay"
        round_number   : int  -- sequential number within the current session
        emotion        : dict -- optional snapshot from EmotionTracker.get_round_snapshot()
        reaction_ms    : int  -- optional time in ms the player took to throw
        """
        # Skip anonymous sessions — there's no profile to save to.
        if not player_name or not player_name.strip():
            return

        profile = self.get_or_create_profile(player_name)

        # Build the core record for this round.
        round_data = {
            "timestamp":      datetime.now().isoformat(),
            "player_gesture": player_gesture,
            "robot_gesture":  robot_gesture,
            "outcome":        outcome,
            "game_mode":      game_mode,
            "round_number":   round_number,
        }

        # Only store reaction time if it was actually measured.
        if reaction_ms is not None:
            round_data["reaction_ms"] = reaction_ms

        # Attach emotion fields if the emotion tracker gave us data.
        if emotion and isinstance(emotion, dict):
            round_data["emotion"]             = emotion.get("emotion",             "Unknown")
            round_data["emotion_confidence"]  = emotion.get("emotion_confidence",  0.0)
            round_data["smile_score"]         = emotion.get("smile_score",         0.0)
            round_data["surprise_score"]      = emotion.get("surprise_score",      0.0)
            round_data["frustration_score"]   = emotion.get("frustration_score",   0.0)

        # Work out how the player moved relative to their previous gesture.
        # "stay" = same, "upgrade" = moved up the cycle, "downgrade" = moved down.
        if profile["rounds"]:
            prev         = profile["rounds"][-1]
            prev_gesture = prev["player_gesture"]

            if player_gesture == prev_gesture:
                round_data["response_type"] = "stay"
            elif UPGRADE.get(prev_gesture) == player_gesture:
                round_data["response_type"] = "upgrade"
            elif DOWNGRADE.get(prev_gesture) == player_gesture:
                round_data["response_type"] = "downgrade"
            else:
                # Any other transition (e.g. Rock -> Scissors) is "lateral".
                round_data["response_type"] = "lateral"

            round_data["previous_gesture"] = prev_gesture
            round_data["previous_outcome"] = prev["outcome"]
        else:
            # This is the player's very first recorded round — no prior to compare.
            round_data["response_type"]    = "first"
            round_data["previous_gesture"] = None
            round_data["previous_outcome"] = None

        profile["rounds"].append(round_data)
        profile["last_played"] = datetime.now().isoformat()

        # Persist to disk and append to the combined Excel log.
        self._save_profile(player_name, profile)
        self._log_to_excel(player_name, round_data)

    # ------------------------------------------------------------------
    # Pattern analysis
    # ------------------------------------------------------------------

    def build_pattern_tables(self, name):
        """
        Compute statistical tables that describe a player's tendencies.

        These tables are what the clone AI reads to pick its next move —
        it samples from the probability distributions to mimic the player.

        Returns a dict with:
            gesture_freq       -- {Rock: 0.4, Paper: 0.35, Scissors: 0.25}
            outcome_response   -- {win: {stay: 0.6, upgrade: 0.2, downgrade: 0.2}, ...}
            transition         -- {Rock: {Rock: 0.3, Paper: 0.5, Scissors: 0.2}, ...}
            outcome_transition -- {win: {Rock: {Rock: ..., ...}, ...}, ...}
            round_count        -- total rounds analysed

        Returns None if the player has no data.
        """
        profile = self.load_profile(name)
        if profile is None or not profile["rounds"]:
            return None

        rounds = profile["rounds"]

        # --- Overall gesture frequency ---
        # Count raw throws then normalise to fractions that sum to 1.
        gesture_counts = {g: 0 for g in GESTURES}
        for r in rounds:
            if r["player_gesture"] in gesture_counts:
                gesture_counts[r["player_gesture"]] += 1

        total       = max(sum(gesture_counts.values()), 1)
        gesture_freq = {g: c / total for g, c in gesture_counts.items()}

        # --- Outcome-conditioned response type ---
        # After a win/loss/draw, how often does the player stay / upgrade / downgrade?
        outcome_response = {o: {"stay": 0, "upgrade": 0, "downgrade": 0} for o in OUTCOMES}
        for r in rounds:
            rt = r.get("response_type")
            po = r.get("previous_outcome")
            if rt in ("stay", "upgrade", "downgrade") and po in OUTCOMES:
                outcome_response[po][rt] += 1

        # Normalise each outcome row to probabilities.
        for o in OUTCOMES:
            total_r = max(sum(outcome_response[o].values()), 1)
            outcome_response[o] = {k: v / total_r for k, v in outcome_response[o].items()}

        # --- Direct move transition matrix ---
        # After throwing gesture X, how likely is the player to throw gesture Y next?
        transition = {g: {g2: 0 for g2 in GESTURES} for g in GESTURES}
        for i in range(len(rounds) - 1):
            curr_g = rounds[i]["player_gesture"]
            next_g = rounds[i + 1]["player_gesture"]
            if curr_g in GESTURES and next_g in GESTURES:
                transition[curr_g][next_g] += 1

        # Normalise each row to probabilities.
        for g in GESTURES:
            total_t = max(sum(transition[g].values()), 1)
            transition[g] = {g2: c / total_t for g2, c in transition[g].items()}

        # --- Outcome + gesture -> next gesture ---
        # More specific version: given the outcome AND what they just played, what's next?
        outcome_transition = {
            o: {g: {g2: 0 for g2 in GESTURES} for g in GESTURES} for o in OUTCOMES
        }
        for i in range(len(rounds) - 1):
            curr_g = rounds[i]["player_gesture"]
            curr_o = rounds[i]["outcome"]
            next_g = rounds[i + 1]["player_gesture"]
            if curr_g in GESTURES and curr_o in OUTCOMES and next_g in GESTURES:
                outcome_transition[curr_o][curr_g][next_g] += 1

        # Normalise each (outcome, gesture) row.
        for o in OUTCOMES:
            for g in GESTURES:
                total_ot = max(sum(outcome_transition[o][g].values()), 1)
                outcome_transition[o][g] = {
                    g2: c / total_ot for g2, c in outcome_transition[o][g].items()
                }

        return {
            "player_name":        name,
            "round_count":        len(rounds),
            "gesture_freq":       gesture_freq,
            "outcome_response":   outcome_response,
            "transition":         transition,
            "outcome_transition": outcome_transition,
        }

    # ------------------------------------------------------------------
    # Saving profiles
    # ------------------------------------------------------------------

    def _save_profile(self, name, profile):
        """Write the profile dict to disk as pretty-printed JSON (2-space indent)."""
        path = self._profile_path(name)
        try:
            with open(path, "w") as f:
                json.dump(profile, f, indent=2)
        except Exception as exc:
            print(f"[ProfileStore] Save error: {exc}")

    # ------------------------------------------------------------------
    # AI state persistence
    # ------------------------------------------------------------------

    def save_ai_state(self, player_name, ai):
        """
        Persist the AI's learned bandit weights alongside the player profile.

        This lets the AI pick up where it left off in the next session instead of
        starting fresh every time. Called when the game ends or the player quits.
        """
        if not player_name or not player_name.strip():
            return
        try:
            profile = self.get_or_create_profile(player_name)
            profile["ai_state"] = {
                "bandit":             getattr(ai, "_bandit", {}),
                "consecutive_wins":   getattr(ai, "_consecutive_wins", 0),
                "consecutive_losses": getattr(ai, "_consecutive_losses", 0),
                "saved_at":           datetime.now().isoformat(),
            }
            self._save_profile(player_name, profile)
        except Exception as exc:
            print(f"[ProfileStore] AI state save error: {exc}")

    def load_ai_state(self, player_name, ai):
        """
        Restore persisted bandit weights into an AI instance.

        We don't blindly replace the weights — saved history is blended in at 70%
        with the fresh (1.0, 1.0) prior at 30%. This means the AI retains what it
        learned but doesn't get permanently locked into stale patterns from old sessions.

        Returns True if state was loaded successfully, False otherwise.
        """
        if not player_name or not player_name.strip():
            return False
        try:
            profile = self.load_profile(player_name)
            if not profile:
                return False

            ai_state = profile.get("ai_state")
            if not ai_state:
                return False

            # Blend in the saved bandit weights.
            if hasattr(ai, "_bandit") and ai_state.get("bandit"):
                for layer, saved in ai_state["bandit"].items():
                    if layer in ai._bandit and isinstance(saved, list) and len(saved) == 2:
                        # 70% of the saved history, floored at 1.0 to keep the prior valid.
                        ai._bandit[layer][0] = max(1.0, saved[0] * 0.7)
                        ai._bandit[layer][1] = max(1.0, saved[1] * 0.7)

            print(f"[ProfileStore] Loaded AI state for {player_name}")
            return True
        except Exception as exc:
            print(f"[ProfileStore] AI state load error: {exc}")
            return False

    # ------------------------------------------------------------------
    # Excel logging
    # ------------------------------------------------------------------

    def _log_to_excel(self, player_name, round_data):
        """
        Append one round's data to the combined research Excel workbook.

        Silently skips if openpyxl isn't installed or the write fails —
        JSON is the authoritative store; Excel is supplementary.
        """
        if not _HAS_OPENPYXL:
            return
        try:
            wb = None

            # Try to open the existing workbook. If it's corrupted, back it up and start fresh.
            if self.excel_path.exists():
                try:
                    wb = load_workbook(self.excel_path)
                except Exception:
                    backup = self.excel_path.with_suffix(".corrupted.xlsx")
                    try:
                        self.excel_path.rename(backup)
                        print(f"[ProfileStore] Corrupted Excel backed up to {backup.name}, creating fresh file.")
                    except Exception:
                        self.excel_path.unlink(missing_ok=True)
                    wb = None

            # No existing workbook (first run, or just backed up) — create one with headers.
            if wb is None:
                wb = Workbook()
                ws = wb.active
                ws.title = "All_Rounds"
                headers = [
                    "timestamp", "player_name", "player_gesture",
                    "robot_gesture", "outcome", "game_mode",
                    "round_number", "response_type",
                    "previous_gesture", "previous_outcome",
                    "emotion", "emotion_confidence",
                    "smile_score", "surprise_score", "frustration_score",
                ]
                ws.append(headers)
                # Style the header row: dark blue fill, white bold text, centred.
                header_fill = PatternFill("solid", fgColor="1F4E78")
                header_font = Font(color="FFFFFF", bold=True)
                for cell in ws[1]:
                    cell.fill      = header_fill
                    cell.font      = header_font
                    cell.alignment = Alignment(horizontal="center")
                ws.freeze_panes = "A2"

            ws = wb["All_Rounds"]

            # Auto-migration: older workbooks might be missing the emotion columns.
            # Add them on the fly rather than failing or silently losing that data.
            current_headers = [c.value for c in ws[1]]
            if "emotion" not in current_headers:
                col = len(current_headers) + 1
                for i, h in enumerate(["emotion", "emotion_confidence",
                                       "smile_score", "surprise_score", "frustration_score"]):
                    cell = ws.cell(row=1, column=col + i, value=h)
                    cell.fill      = PatternFill("solid", fgColor="1F4E78")
                    cell.font      = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center")

            # Append the actual data row (use empty string as fallback for missing fields).
            ws.append([
                round_data.get("timestamp",           ""),
                player_name,
                round_data.get("player_gesture",      ""),
                round_data.get("robot_gesture",       ""),
                round_data.get("outcome",             ""),
                round_data.get("game_mode",           ""),
                round_data.get("round_number",        0),
                round_data.get("response_type",       ""),
                round_data.get("previous_gesture",    ""),
                round_data.get("previous_outcome",    ""),
                round_data.get("emotion",             ""),
                round_data.get("emotion_confidence",  ""),
                round_data.get("smile_score",         ""),
                round_data.get("surprise_score",      ""),
                round_data.get("frustration_score",   ""),
            ])

            wb.save(self.excel_path)
            wb.close()

        except Exception as exc:
            print(f"[ProfileStore] Excel log error: {exc}")

    # ------------------------------------------------------------------
    # Per-player Excel report sheets
    # ------------------------------------------------------------------

    def generate_all_player_reports(self):
        """
        Generate or refresh per-player analysis tabs in the research Excel workbook.

        Each player with at least 5 rounds gets their own sheet showing their
        gesture breakdown, response patterns, and strategic tendencies.
        Returns the number of sheets updated.
        """
        if not _HAS_OPENPYXL:
            return 0
        try:
            # Load the existing workbook, or create a fresh one if none exists yet.
            if self.excel_path.exists():
                wb = load_workbook(self.excel_path)
            else:
                wb = Workbook()
                wb.active.title = "All_Rounds"

            players = self.list_players()
            updated = 0

            for name, count in players:
                # Skip players who don't have enough data for a meaningful analysis.
                if count < 5:
                    continue

                tables = self.build_pattern_tables(name)
                if tables is None:
                    continue

                # Excel sheet names are limited to 31 characters.
                sheet_name = name[:28]
                # Delete the old sheet first so we always get a clean, fresh version.
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]

                ws = wb.create_sheet(sheet_name)
                self._write_player_sheet(ws, name, tables)
                updated += 1

            wb.save(self.excel_path)
            wb.close()
            print(f"[ProfileStore] Updated {updated} player report sheets")
            return updated

        except Exception as exc:
            print(f"[ProfileStore] Report generation error: {exc}")
            return 0

    def _write_player_sheet(self, ws, name, tables):
        """
        Write a single player's analysis to an Excel worksheet.

        Sections written in order:
          1. Title + round count
          2. Overall gesture frequency table
          3. How to beat this player (strategic summary)
          4. Unique traits (human-readable tendencies)
          5. Response patterns (stay/upgrade/downgrade after win/loss/draw)
          6. Move transition matrix (after X, plays Y)
        """
        # Style objects reused throughout the sheet.
        header_fill  = PatternFill("solid", fgColor="1F4E78")
        header_font  = Font(color="FFFFFF", bold=True, size=12)
        section_font = Font(bold=True, size=11)

        row = 1

        # --- Title ---
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value=f"Player Profile: {name}").font = Font(
            bold=True, size=14, color="1F4E78"
        )
        row += 1

        ws.cell(row=row, column=1, value=f"Total rounds: {tables['round_count']}")
        row += 2

        # --- Overall Strategy: gesture frequency table ---
        ws.cell(row=row, column=1, value="OVERALL STRATEGY").font = section_font
        row += 1

        freq      = tables["gesture_freq"]
        favourite = max(freq, key=freq.get)    # most-used gesture
        least     = min(freq, key=freq.get)    # least-used gesture

        # Column headers.
        for col, label in enumerate(["Gesture", "Frequency", "Role"], start=1):
            cell = ws.cell(row=row, column=col, value=label)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # What beats each gesture — used for the "How to Beat" section below.
        counter = {"Rock": "Paper", "Paper": "Scissors", "Scissors": "Rock"}

        for g in GESTURES:
            ws.cell(row=row, column=1, value=g)
            ws.cell(row=row, column=2, value=f"{freq[g]:.0%}")
            # Label the most and least favoured gestures.
            role = "FAVOURITE" if g == favourite else ("Least used" if g == least else "")
            ws.cell(row=row, column=3, value=role)
            row += 1

        row += 1

        # --- How to Beat This Player ---
        ws.cell(row=row, column=1, value="HOW TO BEAT THIS PLAYER").font = section_font
        row += 1

        # Primary advice: counter their favourite gesture.
        best_counter = counter[favourite]
        ws.cell(row=row, column=1, value=f"Primary strategy: Play {best_counter} often")
        ws.cell(row=row, column=2, value=f"Counters their {favourite} ({freq[favourite]:.0%} of throws)")
        row += 1

        # After-loss tendency.
        loss_resp = tables["outcome_response"].get("lose", {})
        loss_max  = max(loss_resp, key=loss_resp.get) if loss_resp else "stay"
        ws.cell(row=row, column=1, value=f"After they lose: they tend to {loss_max}")
        ws.cell(row=row, column=2, value=f"({loss_resp.get(loss_max, 0):.0%} of the time)")
        row += 1

        # After-win tendency.
        win_resp = tables["outcome_response"].get("win", {})
        win_max  = max(win_resp, key=win_resp.get) if win_resp else "stay"
        ws.cell(row=row, column=1, value=f"After they win: they tend to {win_max}")
        ws.cell(row=row, column=2, value=f"({win_resp.get(win_max, 0):.0%} of the time)")
        row += 2

        # --- Unique Traits ---
        ws.cell(row=row, column=1, value="UNIQUE TRAITS").font = section_font
        row += 1
        for trait in self._compute_traits(tables):
            ws.cell(row=row, column=1, value=trait)
            row += 1
        row += 1

        # --- Response Patterns table (after win/loss/draw) ---
        ws.cell(row=row, column=1, value="RESPONSE PATTERNS (after outcome)").font = section_font
        row += 1

        for col, label in enumerate(["After...", "Stay", "Upgrade", "Downgrade"], start=1):
            cell = ws.cell(row=row, column=col, value=label)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")
        row += 1

        for outcome in OUTCOMES:
            resp = tables["outcome_response"].get(outcome, {})
            ws.cell(row=row, column=1, value=outcome.title())
            ws.cell(row=row, column=2, value=f"{resp.get('stay',      0):.0%}")
            ws.cell(row=row, column=3, value=f"{resp.get('upgrade',   0):.0%}")
            ws.cell(row=row, column=4, value=f"{resp.get('downgrade', 0):.0%}")
            row += 1

        row += 1

        # --- Move Transition Matrix ---
        # Rows = gesture they just threw. Columns = gesture they threw next.
        ws.cell(row=row, column=1, value="MOVE TRANSITIONS (after X, plays Y)").font = section_font
        row += 1

        ws.cell(row=row, column=1, value="After...")
        for j, g in enumerate(GESTURES):
            ws.cell(row=row, column=j + 2, value=f"-> {g}")
        for c in range(1, 5):
            cell = ws.cell(row=row, column=c)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")
        row += 1

        trans = tables["transition"]
        for g in GESTURES:
            ws.cell(row=row, column=1, value=g)
            for j, g2 in enumerate(GESTURES):
                ws.cell(row=row, column=j + 2, value=f"{trans[g][g2]:.0%}")
            row += 1

        # Set column widths so content isn't truncated when opened in Excel.
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

    def _compute_traits(self, tables):
        """
        Generate a list of plain-English insight strings about the player's tendencies.

        Each string describes one statistically notable behaviour. These appear in the
        "Unique Traits" section of the Excel sheet and help a human quickly understand
        how to exploit the player.

        Returns a list of strings (always at least one entry).
        """
        traits = []
        freq      = tables["gesture_freq"]
        favourite = max(freq, key=freq.get)

        # --- How dominant is their favourite gesture? ---
        if freq[favourite] > 0.50:
            traits.append(f"Heavy {favourite} player ({freq[favourite]:.0%} of throws)")
        elif freq[favourite] > 0.40:
            traits.append(f"Leans toward {favourite} ({freq[favourite]:.0%} of throws)")
        else:
            traits.append("Relatively balanced  -  no strong favourite")

        # --- Win-stay or win-shift tendency ---
        win_stay = tables["outcome_response"].get("win", {}).get("stay", 0)
        if win_stay > 0.55:
            traits.append(f"Win-stay player  -  repeats winning move {win_stay:.0%} of the time")
        elif win_stay < 0.25:
            traits.append(f"Win-shift player  -  rarely repeats after winning ({win_stay:.0%})")

        # --- Lose-stay or lose-shift tendency ---
        lose_stay = tables["outcome_response"].get("lose", {}).get("stay", 0)
        if lose_stay > 0.45:
            traits.append(f"Stubborn after losses  -  stays with losing move {lose_stay:.0%}")
        elif lose_stay < 0.20:
            traits.append(f"Quick to change after losing  -  only stays {lose_stay:.0%}")

        # --- Do they tend to upgrade after losing? ---
        lose_up = tables["outcome_response"].get("lose", {}).get("upgrade", 0)
        if lose_up > 0.50:
            traits.append(f"Upgrader  -  after losing, upgrades {lose_up:.0%} of the time")

        # --- Find the strongest single gesture-to-gesture transition ---
        trans = tables["transition"]
        best_from, best_to, best_pct = None, None, 0
        for g in GESTURES:
            for g2 in GESTURES:
                if trans[g][g2] > best_pct:
                    best_pct  = trans[g][g2]
                    best_from = g
                    best_to   = g2

        if best_pct > 0.55:
            traits.append(
                f"Predictable sequence: after {best_from}, "
                f"plays {best_to} {best_pct:.0%} of the time"
            )

        # --- How do they tend to respond after a draw? ---
        draw_resp = tables["outcome_response"].get("draw", {})
        draw_max  = max(draw_resp, key=draw_resp.get) if draw_resp else "stay"
        if draw_resp.get(draw_max, 0) > 0.50:
            traits.append(f"After draws, tends to {draw_max} ({draw_resp[draw_max]:.0%})")

        # If nothing notable was found, say so explicitly rather than returning empty.
        if not traits:
            traits.append("No strong identifiable patterns yet")

        return traits

    # ------------------------------------------------------------------
    # CSV export
    # ------------------------------------------------------------------

    def export_csv(self, player_name, output_dir=None):
        """
        Export a player's full round history to a CSV file.

        Useful for loading the data into Excel or running analysis scripts outside the game.
        Returns the file path as a string, or None if the export failed.
        """
        profile = self.load_profile(player_name)
        if not profile:
            return None

        rounds = profile.get("rounds", [])
        if not rounds:
            return None

        # Default output directory: ~/Desktop/CapStone/
        if output_dir is None:
            output_dir = Path.home() / "Desktop" / "CapStone"
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        # Make the filename safe for the filesystem.
        safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in player_name)
        path = output_dir / f"{safe_name}_export.csv"

        fieldnames = [
            "timestamp", "player_gesture", "robot_gesture",
            "outcome", "game_mode", "round_number",
            "response_type", "previous_gesture", "previous_outcome",
            "emotion", "emotion_confidence", "smile_score",
            "surprise_score", "frustration_score",
        ]
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
                writer.writeheader()
                for r in rounds:
                    writer.writerow(r)
            return str(path)
        except Exception as exc:
            print(f"[ProfileStore] CSV export error: {exc}")
            return None

    # ------------------------------------------------------------------
    # Mode-filtered pattern tables
    # ------------------------------------------------------------------

    def build_pattern_tables_filtered(self, name, mode_filter=None):
        """
        Same as build_pattern_tables() but restricted to rounds from one game mode.

        mode_filter:
            None or "All" -> include all rounds (equivalent to the unfiltered version)
            "FairPlay", "Challenge", "Cheat", "Clone" -> only rounds from that mode

        Returns None if there aren't enough matching rounds after filtering.
        """
        profile = self.load_profile(name)
        if profile is None or not profile["rounds"]:
            return None

        all_rounds = profile["rounds"]

        if mode_filter and mode_filter != "All":
            # Normalise some legacy mode name spellings to canonical values.
            _norm_map = {
                "Fair Play":  "FairPlay",
                "fair play":  "FairPlay",
                "Bluff Mode": "Cheat",
            }
            rounds = [
                r for r in all_rounds
                if _norm_map.get(r.get("game_mode", ""), r.get("game_mode", "")) == mode_filter
            ]
        else:
            rounds = all_rounds

        if not rounds:
            return None

        # The computation below is identical to build_pattern_tables(),
        # just operating on the filtered subset instead of all rounds.

        # Overall gesture frequency.
        gesture_counts = {g: 0 for g in GESTURES}
        for r in rounds:
            if r["player_gesture"] in gesture_counts:
                gesture_counts[r["player_gesture"]] += 1
        total        = max(sum(gesture_counts.values()), 1)
        gesture_freq = {g: c / total for g, c in gesture_counts.items()}

        # Outcome-conditioned response type.
        outcome_response = {o: {"stay": 0, "upgrade": 0, "downgrade": 0} for o in OUTCOMES}
        for r in rounds:
            rt = r.get("response_type")
            po = r.get("previous_outcome")
            if rt in ("stay", "upgrade", "downgrade") and po in OUTCOMES:
                outcome_response[po][rt] += 1
        for o in OUTCOMES:
            total_r = max(sum(outcome_response[o].values()), 1)
            outcome_response[o] = {k: v / total_r for k, v in outcome_response[o].items()}

        # Direct move transition matrix.
        transition = {g: {g2: 0 for g2 in GESTURES} for g in GESTURES}
        for i in range(len(rounds) - 1):
            curr_g = rounds[i]["player_gesture"]
            next_g = rounds[i + 1]["player_gesture"]
            if curr_g in GESTURES and next_g in GESTURES:
                transition[curr_g][next_g] += 1
        for g in GESTURES:
            total_t = max(sum(transition[g].values()), 1)
            transition[g] = {g2: c / total_t for g2, c in transition[g].items()}

        # Outcome + gesture -> next gesture.
        outcome_transition = {
            o: {g: {g2: 0 for g2 in GESTURES} for g in GESTURES} for o in OUTCOMES
        }
        for i in range(len(rounds) - 1):
            curr_g = rounds[i]["player_gesture"]
            curr_o = rounds[i]["outcome"]
            next_g = rounds[i + 1]["player_gesture"]
            if curr_g in GESTURES and curr_o in OUTCOMES and next_g in GESTURES:
                outcome_transition[curr_o][curr_g][next_g] += 1
        for o in OUTCOMES:
            for g in GESTURES:
                total_ot = max(sum(outcome_transition[o][g].values()), 1)
                outcome_transition[o][g] = {
                    g2: c / total_ot for g2, c in outcome_transition[o][g].items()
                }

        return {
            "player_name":        name,
            "round_count":        len(rounds),
            "mode_filter":        mode_filter or "All",
            "gesture_freq":       gesture_freq,
            "outcome_response":   outcome_response,
            "transition":         transition,
            "outcome_transition": outcome_transition,
        }

    # ------------------------------------------------------------------
    # Session history
    # ------------------------------------------------------------------

    def get_session_history(self, name, max_sessions=5):
        """
        Group a player's rounds into logical sessions and summarise each one.

        A new session begins when either:
          - The round_number resets (goes lower than the previous round's number),
            which indicates the player started a new game.
          - More than 10 minutes pass between consecutive rounds.

        Returns a list of session summary dicts (up to max_sessions, most recent last):
            date            -- formatted timestamp string, e.g. "14 Mar  18:42"
            mode            -- game mode played in this session
            rounds_played   -- number of rounds in the session
            wins / losses / draws
            win_rate        -- float, 0.0 to 1.0
            avg_reaction_ms -- average throw time in ms, or None if not measured
        """
        profile = self.load_profile(name)
        if profile is None or not profile["rounds"]:
            return []

        rounds   = profile["rounds"]
        sessions = []    # list of lists, each inner list is one session's rounds
        current  = []    # the session we're currently building

        for i, r in enumerate(rounds):
            if i == 0:
                # First round always starts a new session.
                current.append(r)
                continue

            prev = rounds[i - 1]

            # Calculate the time gap between this round and the previous one.
            gap_mins = 0
            try:
                t1       = datetime.fromisoformat(prev.get("timestamp", ""))
                t2       = datetime.fromisoformat(r.get("timestamp",  ""))
                gap_mins = (t2 - t1).total_seconds() / 60
            except Exception:
                pass  # bad or missing timestamps — assume no gap

            # Detect a session boundary: round number reset or long gap.
            rn_reset = r.get("round_number", 0) < prev.get("round_number", 0)
            if rn_reset or gap_mins > 10:
                if current:
                    sessions.append(current)
                current = [r]
            else:
                current.append(r)

        # Don't forget to close off the last open session.
        if current:
            sessions.append(current)

        # Build summary dicts for the last max_sessions sessions.
        result = []
        for sess in sessions[-max_sessions:]:
            wins   = sum(1 for r in sess if r.get("outcome") == "win")
            losses = sum(1 for r in sess if r.get("outcome") == "lose")
            draws  = sum(1 for r in sess if r.get("outcome") == "draw")
            total  = max(len(sess), 1)
            mode   = sess[0].get("game_mode", "?")
            ts     = sess[0].get("timestamp", "")

            try:
                date_str = datetime.fromisoformat(ts).strftime("%d %b  %H:%M")
            except Exception:
                date_str = ts[:16] if ts else "Unknown"

            # Average reaction time — only include values under 3 seconds.
            # Anything higher is likely a pause (looking away, etc.), not a real reaction.
            rt_vals = [
                r.get("reaction_ms") for r in sess
                if r.get("reaction_ms") and r["reaction_ms"] < 3000
            ]
            avg_rt = round(sum(rt_vals) / len(rt_vals)) if rt_vals else None

            result.append({
                "date":            date_str,
                "mode":            mode,
                "rounds_played":   len(sess),
                "wins":            wins,
                "losses":          losses,
                "draws":           draws,
                "win_rate":        wins / total,
                "avg_reaction_ms": avg_rt,
            })

        return result
