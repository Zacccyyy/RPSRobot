"""
simulation_mode.py
==================
Runs headless (no camera, no UI) games at high speed to generate
training data and compare AI strategies against simulated player types.

Why this exists:
    Rather than waiting for real humans to play thousands of rounds, we can
    simulate them quickly to see which AI strategies work best, which player
    styles are hardest to beat, and to generate data for the ML training script.

Usage:
    python simulation_mode.py

Simulated player strategies:
    - "random"        pure random (baseline, no pattern)
    - "win_stay"      win-stay / lose-shift (very common human bias)
    - "cycler"        Rock -> Paper -> Scissors -> repeat
    - "rock_heavy"    60% Rock, 20% Paper, 20% Scissors
    - "anti_pattern"  tries to counter whatever the AI played last
    - "mixed_human"   blend of all the above (most realistic)

AI opponents:
    - "random"        pure random (baseline)
    - "fair_play"     FairPlayAI (heuristic, beatable)
    - "challenge"     ChallengeAI (heuristic, escalating difficulty)
    - "ml"            MLPredictionAI (if a trained model file exists)

Output:
    ~/Desktop/CapStone/simulation_results.xlsx
    (kept separate from real gameplay data)
"""

import random
import time
from pathlib import Path
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from fair_play_ai import FairPlayAI, VALID_GESTURES, COUNTER_MOVE, UPGRADE_MOVE, DOWNGRADE_MOVE


# ==================================================================
# CONFIG — tweak these to change how the simulation runs
# ==================================================================

# Where the Excel results file gets written.
OUTPUT_PATH = Path.home() / "Desktop" / "CapStone" / "simulation_results.xlsx"

# How many independent runs to do for each (player strategy, AI) combination.
# More runs = more reliable averages, but takes longer.
RUNS_PER_COMBO = 10

# How many rounds each run lasts.
ROUNDS_PER_RUN = 100

# Which player strategies and AI opponents to include in the simulation.
PLAYER_STRATEGIES = [
    "random",
    "win_stay",
    "cycler",
    "rock_heavy",
    "anti_pattern",
    "mixed_human",
]

AI_OPPONENTS = [
    "random",
    "fair_play",
    "challenge",
]


# ==================================================================
# Simulated player strategies
# ==================================================================

class SimulatedPlayer:
    """
    Generates moves for one simulated player according to a named strategy.

    Tracks its own history (list of gestures thrown) and a cycle index
    so stateful strategies like "cycler" and "win_stay" work correctly.
    """

    def __init__(self, strategy="random"):
        self.strategy    = strategy
        self.history     = []    # all gestures thrown so far this run
        self.cycle_index = 0     # current position in the Rock->Paper->Scissors cycle

    def reset(self):
        """Clear state so this player can be reused for a new run."""
        self.history     = []
        self.cycle_index = 0

    def choose_move(self, last_outcome=None, last_own_move=None, last_opponent_move=None):
        """
        Pick the next move according to the player's strategy.

        Parameters:
            last_outcome       -- "win", "lose", "draw", or None (first round)
            last_own_move      -- the gesture this player threw last round, or None
            last_opponent_move -- the AI's gesture last round, or None

        Appends the chosen move to self.history and returns it.
        """
        # Dispatch to the correct strategy method.
        if self.strategy == "random":
            move = random.choice(VALID_GESTURES)
        elif self.strategy == "win_stay":
            move = self._win_stay(last_outcome, last_own_move)
        elif self.strategy == "cycler":
            move = self._cycler()
        elif self.strategy == "rock_heavy":
            move = self._rock_heavy()
        elif self.strategy == "anti_pattern":
            move = self._anti_pattern(last_opponent_move)
        elif self.strategy == "mixed_human":
            move = self._mixed_human(last_outcome, last_own_move, last_opponent_move)
        else:
            # Unknown strategy — fall back to random.
            move = random.choice(VALID_GESTURES)

        self.history.append(move)
        return move

    def _win_stay(self, last_outcome, last_own_move):
        """
        Win-stay / lose-shift bias: a very common human pattern.

        After a win:  80% chance of repeating the same gesture, 20% random.
        After a loss: 70% chance of switching to a different gesture, 30% stay.
        After a draw: 50/50 stay or switch.
        First round:  pure random (no prior move to base a decision on).
        """
        if last_outcome is None or last_own_move is None:
            return random.choice(VALID_GESTURES)

        if last_outcome == "win":
            # Usually stick with the winning move.
            return last_own_move if random.random() < 0.80 else random.choice(VALID_GESTURES)

        if last_outcome == "lose":
            # Usually switch after losing.
            if random.random() < 0.70:
                others = [g for g in VALID_GESTURES if g != last_own_move]
                return random.choice(others)
            return last_own_move

        # Draw: flip a coin on whether to stay or switch.
        return last_own_move if random.random() < 0.50 else random.choice(VALID_GESTURES)

    def _cycler(self):
        """
        Cycle through Rock -> Paper -> Scissors in order, repeating forever.
        Uses self.cycle_index to track the current position.
        """
        cycle = ["Rock", "Paper", "Scissors"]
        move = cycle[self.cycle_index % 3]
        self.cycle_index += 1
        return move

    def _rock_heavy(self):
        """
        Throw Rock 60% of the time, Paper 20%, Scissors 20%.
        Simulates a player who has a strong Rock preference.
        """
        r = random.random()
        if r < 0.60:
            return "Rock"
        if r < 0.80:
            return "Paper"
        return "Scissors"

    def _anti_pattern(self, last_opponent_move):
        """
        Try to counter whatever the AI played last round.
        Adds 35% random noise so it's not perfectly predictable.
        First round is pure random (no opponent move to react to yet).
        """
        if last_opponent_move is None:
            return random.choice(VALID_GESTURES)

        # 65% of the time: counter the AI's last move. 35%: throw randomly.
        if random.random() < 0.65:
            return COUNTER_MOVE[last_opponent_move]
        return random.choice(VALID_GESTURES)

    def _mixed_human(self, last_outcome, last_own_move, last_opponent_move):
        """
        A realistic blend of several human tendencies, sampled each round.

        Weights are chosen to reflect how common each bias is in real players:
          40%: win-stay / lose-shift
          20%: anti-pattern (counter the AI)
          15%: rock-heavy
          10%: upgrade after a loss
          15%: pure random
        """
        r = random.random()

        if r < 0.40:
            return self._win_stay(last_outcome, last_own_move)
        if r < 0.60:
            return self._anti_pattern(last_opponent_move)
        if r < 0.75:
            return self._rock_heavy()
        if r < 0.85:
            # After losing, players often "upgrade" — step forward in the cycle.
            if last_outcome == "lose" and last_own_move is not None:
                return UPGRADE_MOVE[last_own_move]
            return random.choice(VALID_GESTURES)
        return random.choice(VALID_GESTURES)


# ==================================================================
# AI opponent factory
# ==================================================================

def create_ai_opponent(ai_type):
    """
    Create an AI opponent and return a callable that produces its move.

    Returns a (ai_instance, get_move_fn) tuple.
    get_move_fn signature: (history, streak, round_number) -> gesture_string

    Returns (None, None) if the AI type isn't recognised or its module is missing.
    """
    if ai_type == "random":
        # Random baseline: ignore all arguments and pick uniformly.
        return None, lambda history, streak, rn: random.choice(VALID_GESTURES)

    if ai_type == "fair_play":
        ai = FairPlayAI()
        def get_move(history, streak, round_number):
            return ai.choose_robot_move(history=history, round_number=round_number)
        return ai, get_move

    if ai_type == "challenge":
        # ChallengeAI is a subclass of FairPlayAI that also takes a streak argument.
        from challenge_ai import ChallengeAI
        ai = ChallengeAI()
        def get_move(history, streak, round_number):
            return ai.choose_robot_move(history=history, streak=streak, round_number=round_number)
        return ai, get_move

    if ai_type == "ml":
        # ML model is optional — skip gracefully if the file or module doesn't exist.
        try:
            from ml_model import MLPredictionAI
            model_path = Path.home() / "Desktop" / "CapStone" / "rps_ml_model.pkl"
            if not model_path.exists():
                print(f"[Simulation] ML model not found at {model_path}, skipping.")
                return None, None
            ai = MLPredictionAI(model_path=str(model_path))
            def get_move(history, streak, round_number):
                return ai.choose_robot_move(history=history, streak=streak, round_number=round_number)
            return ai, get_move
        except ImportError:
            print("[Simulation] ML model imports failed, skipping.")
            return None, None

    # Unknown type — caller handles the (None, None) case.
    return None, None


# ==================================================================
# RPS outcome resolution
# ==================================================================

# What each gesture beats — used by compare_rps() below.
BEATS = {
    "Rock":     "Scissors",
    "Paper":    "Rock",
    "Scissors": "Paper",
}


def compare_rps(player_move, robot_move):
    """
    Resolve a single round of Rock-Paper-Scissors from the player's perspective.
    Returns "win", "lose", or "draw".
    """
    if player_move == robot_move:
        return "draw"
    if BEATS[player_move] == robot_move:
        return "win"
    return "lose"


# ==================================================================
# Single simulation run
# ==================================================================

def run_single_game(player_strategy, ai_type, num_rounds):
    """
    Simulate one complete game between a player strategy and an AI opponent.

    Returns a dict:
        {
            "rounds":      [list of round dicts],
            "final_streak": int,   # longest winning streak the player achieved
            "player_wins": int,
            "robot_wins":  int,
            "draws":       int,
        }
    Returns None if the AI type doesn't exist or couldn't be loaded.
    """
    player = SimulatedPlayer(strategy=player_strategy)
    ai_instance, ai_get_move = create_ai_opponent(ai_type)

    # Couldn't create the AI (missing module or unknown type) — bail out.
    if ai_get_move is None:
        return None

    # Reset AI learned state at the start of each fresh game.
    if ai_instance is not None and hasattr(ai_instance, "reset"):
        ai_instance.reset()

    history      = []
    streak       = 0      # current player win streak (resets on loss or draw)
    player_wins  = 0
    robot_wins   = 0
    draws        = 0

    # Track previous round values so strategies like win_stay can read them.
    last_outcome      = None
    last_player_move  = None
    last_robot_move   = None

    for round_num in range(1, num_rounds + 1):

        # Player picks their gesture for this round.
        player_move = player.choose_move(
            last_outcome=last_outcome,
            last_own_move=last_player_move,
            last_opponent_move=last_robot_move,
        )

        # AI picks its gesture (it reads the shared history list).
        robot_move = ai_get_move(history, streak, round_num)

        # Determine who won this round.
        outcome = compare_rps(player_move, robot_move)

        if outcome == "win":
            player_wins += 1
            streak += 1
            player_outcome = "win"
        elif outcome == "lose":
            robot_wins += 1
            streak = 0
            player_outcome = "lose"
        else:
            draws += 1
            streak = 0
            player_outcome = "draw"

        # Classify how the player moved relative to their previous gesture.
        response_type = None
        if last_player_move is not None:
            if player_move == last_player_move:
                response_type = "stay"
            elif UPGRADE_MOVE.get(last_player_move) == player_move:
                response_type = "upgrade"
            elif DOWNGRADE_MOVE.get(last_player_move) == player_move:
                response_type = "downgrade"

        # Build the round record — same structure the real game produces.
        round_record = {
            "round_number":           round_num,
            "player_gesture":         player_move,
            "robot_gesture":          robot_move,
            "player_outcome":         player_outcome,
            "previous_player_gesture": last_player_move,
            "player_response_type":   response_type,
        }
        history.append(round_record)

        # Update trackers for the next iteration.
        last_outcome     = player_outcome
        last_player_move = player_move
        last_robot_move  = robot_move

    # Calculate the longest consecutive win streak the player had during this game.
    max_streak = 0
    current    = 0
    for r in history:
        if r["player_outcome"] == "win":
            current   += 1
            max_streak = max(max_streak, current)
        else:
            current = 0

    return {
        "rounds":       history,
        "final_streak": max_streak,
        "player_wins":  player_wins,
        "robot_wins":   robot_wins,
        "draws":        draws,
    }


# ==================================================================
# Excel output
# ==================================================================

def save_results_to_excel(all_results, output_path):
    """
    Save simulation results to an Excel workbook with two sheets:
        - Sim_Summary  — one row per run (aggregated stats)
        - Sim_Rounds   — every individual round (raw data)

    all_results is a dict keyed by (strategy, ai_type) tuples,
    with a list of run-result dicts as the value.
    """
    wb = Workbook()

    # --- Summary sheet: one row per run ---
    summary_ws = wb.active
    summary_ws.title = "Sim_Summary"
    summary_ws.append([
        "run_id", "player_strategy", "ai_opponent", "rounds_played",
        "player_wins", "robot_wins", "draws",
        "player_win_rate", "robot_win_rate", "max_streak", "timestamp",
    ])

    # --- Rounds sheet: every individual round ---
    rounds_ws = wb.create_sheet("Sim_Rounds")
    rounds_ws.append([
        "run_id", "player_strategy", "ai_opponent", "round_number",
        "player_gesture", "robot_gesture", "round_result",
        "player_outcome", "previous_player_gesture", "player_response_type",
    ])

    timestamp   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    run_counter = 0

    # Loop through every (strategy, ai) combination and each run within it.
    for (strategy, ai_type), runs in all_results.items():
        for run in runs:
            run_counter += 1
            run_id = f"SIM-{run_counter:04d}"

            # Win rates are calculated excluding draws (they don't count as decided rounds).
            total_decided = run["player_wins"] + run["robot_wins"]
            player_wr = run["player_wins"] / total_decided if total_decided > 0 else 0.0
            robot_wr  = run["robot_wins"]  / total_decided if total_decided > 0 else 0.0

            summary_ws.append([
                run_id,
                strategy,
                ai_type,
                len(run["rounds"]),
                run["player_wins"],
                run["robot_wins"],
                run["draws"],
                round(player_wr, 4),
                round(robot_wr, 4),
                run["final_streak"],
                timestamp,
            ])

            # Write every round into the rounds sheet, translating outcome to a readable label.
            outcome_label = {"win": "player_win", "lose": "robot_win", "draw": "draw"}
            for r in run["rounds"]:
                rounds_ws.append([
                    run_id,
                    strategy,
                    ai_type,
                    r["round_number"],
                    r["player_gesture"],
                    r["robot_gesture"],
                    outcome_label.get(r["player_outcome"], r["player_outcome"]),
                    r["player_outcome"],
                    r["previous_player_gesture"],
                    r["player_response_type"],
                ])

    # --- Header row formatting: dark blue background, white bold text ---
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in [summary_ws, rounds_ws]:
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"   # freeze the header so it stays visible when scrolling

    # Set column widths so the data isn't truncated in Excel.
    col_widths_summary = {
        "A": 14, "B": 18, "C": 14, "D": 14, "E": 14,
        "F": 14, "G": 10, "H": 16, "I": 16, "J": 14, "K": 22,
    }
    col_widths_rounds = {
        "A": 14, "B": 18, "C": 14, "D": 14, "E": 16,
        "F": 16, "G": 14, "H": 16, "I": 22, "J": 20,
    }
    for col, width in col_widths_summary.items():
        summary_ws.column_dimensions[col].width = width
    for col, width in col_widths_rounds.items():
        rounds_ws.column_dimensions[col].width = width

    # Save the workbook — catch PermissionError in case Excel has it open.
    try:
        wb.save(output_path)
        print(f"\nResults saved to: {output_path}")
    except PermissionError:
        print(f"\nCould not save — please close {output_path} in Excel and try again.")
    finally:
        wb.close()


# ==================================================================
# Main simulation runner — called from main.py, GUI, or CLI
# ==================================================================

def run_simulation(
    player_strategies=None,
    ai_opponents=None,
    runs_per_combo=RUNS_PER_COMBO,
    rounds_per_run=ROUNDS_PER_RUN,
    save_excel=True,
    output_path=None,
):
    """
    Run the full simulation across all (player strategy, AI) combinations.

    All parameters have defaults from the CONFIG section at the top of this file,
    but can be overridden when calling from code.

    Returns a summary dict:
        {
            "elapsed_seconds": float,
            "total_rounds":    int,
            "total_runs":      int,
            "combo_results":   [
                {
                    "strategy":       str,
                    "ai":             str,
                    "player_win_rate": float,
                    "robot_win_rate":  float,
                    "draw_rate":       float,
                    "avg_streak":      float,
                    "runs":            int,
                },
                ...
            ],
            "best_ai":         str,   # hardest AI for players to beat
            "worst_ai":        str,   # easiest AI for players to beat
            "best_strategy":   str,   # most effective player strategy
            "worst_strategy":  str,   # least effective player strategy
        }
    """
    # Apply defaults for anything not supplied by the caller.
    if player_strategies is None:
        player_strategies = PLAYER_STRATEGIES
    if ai_opponents is None:
        ai_opponents = AI_OPPONENTS
    if output_path is None:
        output_path = OUTPUT_PATH

    all_results      = {}   # raw run data, keyed by (strategy, ai_type)
    combo_summaries  = []   # aggregated stats per combination
    start_time       = time.time()

    # Run every combination of player strategy and AI opponent.
    for strategy in player_strategies:
        for ai_type in ai_opponents:
            runs = []

            # Repeat each combination multiple times for reliable averages.
            for _ in range(runs_per_combo):
                result = run_single_game(strategy, ai_type, rounds_per_run)
                if result is not None:
                    runs.append(result)

            all_results[(strategy, ai_type)] = runs

            # Aggregate the results across all runs for this combination.
            if runs:
                total_pw  = sum(r["player_wins"] for r in runs)
                total_rw  = sum(r["robot_wins"]  for r in runs)
                total_d   = sum(r["draws"]        for r in runs)
                total_all = total_pw + total_rw + total_d
                avg_streak = sum(r["final_streak"] for r in runs) / len(runs)

                combo_summaries.append({
                    "strategy":        strategy,
                    "ai":              ai_type,
                    "player_win_rate": total_pw / total_all if total_all > 0 else 0.0,
                    "robot_win_rate":  total_rw / total_all if total_all > 0 else 0.0,
                    "draw_rate":       total_d  / total_all if total_all > 0 else 0.0,
                    "avg_streak":      round(avg_streak, 1),
                    "runs":            len(runs),
                })

    elapsed = time.time() - start_time

    # Count totals across everything.
    total_runs   = sum(len(r) for r in all_results.values())
    total_rounds = sum(len(r["rounds"]) for runs in all_results.values() for r in runs)

    # --- Find the best and worst AI and player strategy ---
    # Group win rates by AI and by strategy, then average each group.
    ai_win_rates       = {}
    strategy_win_rates = {}
    for s in combo_summaries:
        ai_win_rates.setdefault(s["ai"], []).append(s["robot_win_rate"])
        strategy_win_rates.setdefault(s["strategy"], []).append(s["player_win_rate"])

    def _avg(lst):
        """Simple average of a list. Returns 0.0 for an empty list."""
        return sum(lst) / len(lst) if lst else 0.0

    # Best AI = highest robot win rate across all player strategies.
    best_ai      = max(ai_win_rates,       key=lambda k: _avg(ai_win_rates[k]))       if ai_win_rates       else "N/A"
    worst_ai     = min(ai_win_rates,       key=lambda k: _avg(ai_win_rates[k]))       if ai_win_rates       else "N/A"
    # Best strategy = highest player win rate across all AIs.
    best_strategy  = max(strategy_win_rates, key=lambda k: _avg(strategy_win_rates[k])) if strategy_win_rates else "N/A"
    worst_strategy = min(strategy_win_rates, key=lambda k: _avg(strategy_win_rates[k])) if strategy_win_rates else "N/A"

    # Optionally write everything to the Excel file.
    if save_excel:
        save_results_to_excel(all_results, output_path)

    return {
        "elapsed_seconds": round(elapsed, 1),
        "total_rounds":    total_rounds,
        "total_runs":      total_runs,
        "combo_results":   combo_summaries,
        "best_ai":         best_ai,
        "worst_ai":        worst_ai,
        "best_strategy":   best_strategy,
        "worst_strategy":  worst_strategy,
    }


# ==================================================================
# CLI entry point
# ==================================================================

def main():
    """
    Run the simulation from the command line and print a summary to stdout.
    Called automatically when this file is run directly.
    """
    print("=" * 60)
    print("RPS Simulation Mode")
    print("=" * 60)
    print()
    print(f"Runs per combination:   {RUNS_PER_COMBO}")
    print(f"Rounds per run:         {ROUNDS_PER_RUN}")
    print(f"Player strategies:      {len(PLAYER_STRATEGIES)}")
    print(f"AI opponents:           {len(AI_OPPONENTS)}")

    total_combos = len(PLAYER_STRATEGIES) * len(AI_OPPONENTS)
    total_runs   = total_combos * RUNS_PER_COMBO
    total_rounds = total_runs   * ROUNDS_PER_RUN
    print(f"Total combinations:     {total_combos}")
    print(f"Total runs:             {total_runs}")
    print(f"Total rounds:           {total_rounds:,}")
    print()

    results = run_simulation()

    # Print one line per (strategy, AI) combination.
    for s in results["combo_results"]:
        print(
            f"  {s['strategy']:15s} vs {s['ai']:12s}  |  "
            f"Player WR: {s['player_win_rate']:.1%}  "
            f"Avg streak: {s['avg_streak']:.1f}  "
            f"({s['runs']} runs)"
        )

    print()
    print(f"Simulation complete in {results['elapsed_seconds']}s")
    print(f"Generated {results['total_rounds']:,} rounds across {results['total_runs']} runs.")
    print()
    print(f"Strongest AI:          {results['best_ai']}")
    print(f"Weakest AI:            {results['worst_ai']}")
    print(f"Best player strategy:  {results['best_strategy']}")
    print(f"Worst player strategy: {results['worst_strategy']}")
    print()
    print("You can now re-run ml_training_script.py with simulation data.")


if __name__ == "__main__":
    main()
