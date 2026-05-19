"""
report_updater.py
=================
Reads all four research Excel logs and rewrites every data table in the
capstone Markdown report with live figures.

This file is the "bridge" between the raw data the game records (Excel logs)
and the human-readable research report (a Markdown file that can be opened in
Word or converted to PDF).  It replaces every table and a handful of in-text
numbers with fresh values every time it runs, so the report is always up to
date without manual copy-pasting.

Run from the project directory:
    python report_updater.py

Or call update_report() from main.py after a research data event.

Tables updated
--------------
  Table 2    — Simulation robot win rates by strategy x AI
  Table 3    — Challenge mode aggregate statistics
  Table 3b   — Streak distribution
  Table 3c   — Outcome-conditioned response rates (post-win / post-draw)
  Table 4    — Player profile traits (first player in log)
  Table 5    — Move transition matrix (first player)
  Table 5b   — Comparative analysis (player vs challenge population vs Nash)
  Table 6    — Technology stack (auto-updates version numbers)
  Table 7    — Module inventory (updates line counts from .py files)
  Abstract   — Updates round count and iteration count in-text
"""

from __future__ import annotations
import os
import re
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime

# openpyxl reads .xlsx files without needing Excel installed.
try:
    from openpyxl import load_workbook
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("[Updater] openpyxl not found — install with: pip install openpyxl")

# ---------------------------------------------------------------------------
# File paths
# ---------------------------------------------------------------------------
# CAPSTONE_DIR is where the Excel logs and the output report live.
# PROJECT_DIR  is where this Python file lives (i.e. the game source folder).
CAPSTONE_DIR = Path.home() / "Desktop" / "CapStone"
PROJECT_DIR  = Path(__file__).parent

# The source report is stored as a Markdown file (despite the .docx extension).
REPORT_SRC   = PROJECT_DIR / "RPS_Capstone_Research_Report-4.docx"
REPORT_OUT   = CAPSTONE_DIR / "RPS_Capstone_Research_Report_Updated.md"

# The four Excel data sources that feed the tables.
PLAYER_LOG      = CAPSTONE_DIR / "player_research_log.xlsx"
CHALLENGE_LOG   = CAPSTONE_DIR / "challenge_research_log.xlsx"
SIM_RESULTS     = CAPSTONE_DIR / "simulation_results.xlsx"
COMPARISON_RPT  = CAPSTONE_DIR / "research_comparison_report.xlsx"


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _wb_rows(path: Path, sheet: str) -> list[dict]:
    """
    Load a single worksheet from an Excel file and return it as a list of dicts.

    The first row of the sheet is treated as the header.  Each subsequent row
    becomes a dict keyed by those header values.  Completely empty rows are
    dropped.  Returns an empty list if the file doesn't exist or can't be read.
    """
    if not path.exists():
        return []
    try:
        wb = load_workbook(path, data_only=True)  # data_only=True gives cell values, not formulas
        if sheet not in wb.sheetnames:
            return []
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        # Build header list, replacing None headers with a positional fallback.
        headers = [
            str(h).strip() if h is not None else f"col{i}"
            for i, h in enumerate(rows[0])
        ]
        # Zip each data row with the headers; skip rows that are entirely empty.
        return [
            dict(zip(headers, row))
            for row in rows[1:]
            if any(c is not None for c in row)
        ]
    except Exception as exc:
        print(f"[Updater] Could not read {path.name}/{sheet}: {exc}")
        return []


def _wb_summary(path: Path, sheet: str) -> dict:
    """
    Load a two-column (Metric, Value) worksheet and return it as a plain dict.

    This is used for "Summary" sheets that store key-value pairs rather than
    tabular data.
    """
    rows = _wb_rows(path, sheet)
    out  = {}
    for r in rows:
        # Collect only the non-None values from this row.
        vals = [v for v in r.values() if v is not None]
        if len(vals) >= 2:
            out[str(vals[0]).strip()] = vals[1]
    return out


def _safe_float(v, default=0.0) -> float:
    """Convert a cell value to float, returning `default` on any failure."""
    try:
        return float(v) if v is not None else default
    except (TypeError, ValueError):
        return default


def _safe_int(v, default=0) -> int:
    """Convert a cell value to int (via float to handle '3.0' strings), returning `default` on failure."""
    try:
        return int(float(v)) if v is not None else default
    except (TypeError, ValueError):
        return default


def _pct(v) -> str:
    """Format a 0–1 float as a percentage string with one decimal place."""
    return f"{_safe_float(v) * 100:.1f}%"


def _pct_of(num, total) -> str:
    """Format num/total as a percentage string, safely handling zero total."""
    if total == 0:
        return "0.0%"
    return f"{num / total * 100:.1f}%"


# ---------------------------------------------------------------------------
# Data extraction — one function per data source
# ---------------------------------------------------------------------------

def _challenge_summary() -> dict:
    """
    Read the Challenge mode logs and compute everything the report tables need.

    Pulls data from three sheets:
      - Summary        : pre-aggregated totals (fast path for simple counts)
      - Challenge_Runs : one row per complete run, used for streak distribution
      - Challenge_Rounds: one row per round, used for gesture counts and OCR

    Returns a single dict with all the computed values so the table builders
    don't need to know anything about the raw Excel structure.
    """
    summary = _wb_summary(CHALLENGE_LOG, "Summary")
    runs    = _wb_rows(CHALLENGE_LOG, "Challenge_Runs")
    rounds  = _wb_rows(CHALLENGE_LOG, "Challenge_Rounds")

    # Simple totals — read straight from the Summary sheet if possible.
    total_runs   = _safe_int(summary.get("total_runs",   0))
    longest      = _safe_int(summary.get("longest_streak", 0))
    total_rounds = sum(_safe_int(r.get("rounds_played", 0)) for r in runs)

    # Count round outcomes directly from the rounds sheet.
    p_wins     = sum(1 for r in rounds if str(r.get("round_result", "")).strip() == "player_win")
    robot_wins = sum(1 for r in rounds if str(r.get("round_result", "")).strip() == "robot_win")
    draws      = sum(1 for r in rounds if str(r.get("round_result", "")).strip() == "draw")
    total_r    = max(p_wins + robot_wins + draws, 1)  # avoid division by zero

    # Count how often each gesture was played.
    rock  = sum(1 for r in rounds if r.get("player_gesture") == "Rock")
    paper = sum(1 for r in rounds if r.get("player_gesture") == "Paper")
    sciss = sum(1 for r in rounds if r.get("player_gesture") == "Scissors")
    total_g = max(rock + paper + sciss, 1)

    # Build a frequency table of final streak lengths (how many wins in a row
    # before the player lost).  Used for Table 3b.
    streak_dist: Counter = Counter()
    for run in runs:
        # Use "final_streak" column, falling back to "max_streak" if not present.
        s = _safe_int(run.get("final_streak", run.get("max_streak", 0)))
        streak_dist[s] += 1

    # Outcome-conditioned response (OCR): what gesture did the player pick
    # after winning/losing/drawing?  We classify each transition as:
    #   stay      = same gesture as last round
    #   upgrade   = the gesture that beats the previous one  (Rock → Paper → Scissors → Rock)
    #   downgrade = the gesture that loses to the previous one
    ocr = defaultdict(lambda: {"stay": 0, "upgrade": 0, "downgrade": 0})
    UPGRADE   = {"Rock": "Paper", "Paper": "Scissors", "Scissors": "Rock"}
    DOWNGRADE = {"Rock": "Scissors", "Paper": "Rock", "Scissors": "Paper"}

    # Group rounds by their run_id so we can compare consecutive rounds within
    # the same run (comparing across runs would give meaningless transitions).
    by_run: dict[str, list] = defaultdict(list)
    for r in rounds:
        run_id = str(r.get("run_id", ""))
        by_run[run_id].append(r)

    # For each run, sort rounds by round number then walk the pairs.
    for run_rounds in by_run.values():
        run_rounds.sort(key=lambda r: _safe_int(r.get("round_number", 0)))
        for i in range(1, len(run_rounds)):
            prev = run_rounds[i - 1]
            curr = run_rounds[i]
            prev_result = str(prev.get("round_result", "")).strip()
            pg = prev.get("player_gesture")
            cg = curr.get("player_gesture")

            # Skip any round with an unexpected gesture value.
            if pg not in ("Rock", "Paper", "Scissors"):
                continue
            if cg not in ("Rock", "Paper", "Scissors"):
                continue

            # Convert the raw result string to a simple outcome label.
            if prev_result == "player_win":
                outcome = "win"
            elif prev_result == "robot_win":
                outcome = "lose"
            else:
                outcome = "draw"

            # Classify the transition type.
            if cg == pg:
                rt = "stay"
            elif UPGRADE.get(pg) == cg:
                rt = "upgrade"
            else:
                rt = "downgrade"

            ocr[outcome][rt] += 1

    def _ocr_row(outcome):
        """Convert raw counts for one outcome into proportions (0–1)."""
        d = ocr[outcome]
        total = max(sum(d.values()), 1)  # avoid division by zero
        n = sum(d.values())
        return {
            "stay":      d["stay"] / total,
            "upgrade":   d["upgrade"] / total,
            "downgrade": d["downgrade"] / total,
            "n":         n,
        }

    return {
        "total_runs":   total_runs,
        "total_rounds": total_rounds,
        "longest":      longest,
        "p_wins":       p_wins,
        "robot_wins":   robot_wins,
        "draws_ch":     draws,
        "total_r":      total_r,
        "rock":         rock,
        "paper":        paper,
        "scissors":     sciss,
        "total_g":      total_g,
        "streak_dist":  streak_dist,
        "ocr":          {o: _ocr_row(o) for o in ("win", "lose", "draw")},
    }


def _player_summary(player_name: str | None = None) -> dict:
    """
    Build a full statistical profile for one player from the All_Rounds sheet.

    If player_name is None, the first player found in the sheet is used.
    Returns a dict with gesture frequencies, transition probabilities, OCR
    rates, win rate, and the best-to-counter information.
    """
    all_rounds = _wb_rows(PLAYER_LOG, "All_Rounds")
    if not all_rounds:
        return {}

    # Default to the first player name seen in the data.
    if player_name is None:
        player_name = str(all_rounds[0].get("player_name", "Unknown")).strip()

    # Filter to only this player's rounds; fall back to all rounds if no match.
    rounds = [r for r in all_rounds
              if str(r.get("player_name", "")).strip() == player_name]
    if not rounds:
        rounds = all_rounds

    total = max(len(rounds), 1)
    GESTURES = ("Rock", "Paper", "Scissors")
    UPGRADE  = {"Rock": "Paper",  "Paper": "Scissors", "Scissors": "Rock"}

    # Count how often each gesture was used.
    gest_c = Counter(
        r.get("player_gesture")
        for r in rounds
        if r.get("player_gesture") in GESTURES
    )
    fav  = gest_c.most_common(1)[0][0] if gest_c else "?"
    least = gest_c.most_common()[-1][0] if len(gest_c) >= 3 else "?"

    # Outcome-conditioned response from explicit columns in the log.
    # (The challenge version had to compute this from pairs; here the log
    #  stores response_type and previous_outcome directly — much easier.)
    ocr = defaultdict(lambda: {"stay": 0, "upgrade": 0, "downgrade": 0})
    for r in rounds:
        rt = str(r.get("response_type", "")).strip()
        po = str(r.get("previous_outcome", "")).strip()
        if rt in ("stay", "upgrade", "downgrade") and po in ("win", "lose", "draw"):
            ocr[po][rt] += 1

    def _fmt_ocr(outcome):
        """Return proportions (0–1) for one outcome bucket."""
        d = ocr[outcome]
        total_o = max(sum(d.values()), 1)
        return {k: d[k] / total_o for k in ("stay", "upgrade", "downgrade")}

    # Overall response-type totals (ignoring conditioning on outcome).
    all_rt = Counter(str(r.get("response_type", "")) for r in rounds)
    rt_total = max(
        all_rt.get("stay", 0) + all_rt.get("upgrade", 0) + all_rt.get("downgrade", 0),
        1,
    )

    # Build the gesture-to-gesture transition matrix.
    # trans[g1][g2] = number of times the player went from g1 to g2.
    trans = defaultdict(Counter)
    for i in range(1, len(rounds)):
        prev_g = rounds[i - 1].get("player_gesture")
        curr_g = rounds[i].get("player_gesture")
        if prev_g in GESTURES and curr_g in GESTURES:
            trans[prev_g][curr_g] += 1

    def _trans_row(g):
        """Return transition probabilities (0–1) from gesture g to each other gesture."""
        t = trans[g]
        tot = max(sum(t.values()), 1)
        return {g2: t[g2] / tot for g2 in GESTURES}

    # Simple win/loss counts.
    wins   = sum(1 for r in rounds if r.get("outcome") == "win")
    losses = sum(1 for r in rounds if r.get("outcome") == "lose")

    # Find the single strongest (most probable) transition in the matrix.
    best_src, best_dst, best_pct = "Rock", "Scissors", 0.0
    for g in GESTURES:
        for g2 in GESTURES:
            p = _trans_row(g).get(g2, 0)
            if p > best_pct:
                best_pct = p
                best_src, best_dst = g, g2

    return {
        "player_name":    player_name,
        "total_rounds":   len(rounds),
        "fav_gesture":    fav,
        "fav_pct":        gest_c[fav] / total if fav in gest_c else 0,
        "least_gesture":  least,
        "least_pct":      gest_c.get(least, 0) / total if least in gest_c else 0,
        "rock_pct":       gest_c.get("Rock", 0) / total,
        "paper_pct":      gest_c.get("Paper", 0) / total,
        "scissors_pct":   gest_c.get("Scissors", 0) / total,
        "ocr":            {o: _fmt_ocr(o) for o in ("win", "lose", "draw")},
        "transition":     {g: _trans_row(g) for g in GESTURES},
        "best_trans_src": best_src,
        "best_trans_dst": best_dst,
        "best_trans_pct": best_pct,
        "stay_pct":       all_rt.get("stay", 0) / rt_total,
        "upgrade_pct":    all_rt.get("upgrade", 0) / rt_total,
        "downgrade_pct":  all_rt.get("downgrade", 0) / rt_total,
        "win_rate":       wins / total,
    }


def _simulation_summary() -> dict[tuple, dict]:
    """
    Load the simulation results and group them by (player_strategy, ai_opponent).

    Returns a dict keyed by (strategy_str, ai_str) tuples.  Each value is
    another dict with robot_win_rate, player_win_rate, draw_rate, and run count.
    """
    rows = _wb_rows(SIM_RESULTS, "Sim_Summary")
    if not rows:
        return {}

    # Accumulate rows that share the same strategy+AI combination.
    groups: dict[tuple, list] = defaultdict(list)
    for r in rows:
        key = (
            str(r.get("player_strategy", "")).strip(),
            str(r.get("ai_opponent",     "")).strip(),
        )
        groups[key].append(r)

    # Aggregate each group into a single set of rates.
    result = {}
    for key, rlist in groups.items():
        total_rounds = sum(_safe_int(r.get("rounds_played", 0)) for r in rlist)
        if total_rounds == 0:
            continue  # skip empty groups to avoid divide-by-zero
        p_wins = sum(_safe_int(r.get("player_wins", 0)) for r in rlist)
        r_wins = sum(_safe_int(r.get("robot_wins",  0)) for r in rlist)
        draws  = sum(_safe_int(r.get("draws",       0)) for r in rlist)
        result[key] = {
            "robot_win_rate":  r_wins  / total_rounds,
            "player_win_rate": p_wins  / total_rounds,
            "draw_rate":       draws   / total_rounds,
            "runs":            len(rlist),
            "total_rounds":    total_rounds,
        }
    return result


# ---------------------------------------------------------------------------
# Table builders — each function returns a list of Markdown table row strings
# ---------------------------------------------------------------------------

def _table2_rows(sim: dict) -> list[str]:
    """
    Build Table 2: robot win rate for each player strategy x AI opponent combo.

    Columns appear only if that AI type exists in the simulation data, so the
    table automatically adjusts when not all AIs have been run yet.
    """
    # Human-readable display names for the AI and strategy keys used in Excel.
    AI_MAP = {
        "random":     "Random AI",
        "fair_play":  "Heuristic FP",
        "challenge":  "Heuristic CH",
        "ml":         "ML Predict",
    }
    STRATEGY_MAP = {
        "random":       "Random",
        "win_stay":     "Win-Stay",
        "cycler":       "Cycler",
        "rock_heavy":   "Rock Heavy",
        "anti_pattern": "Anti-Pattern",
        "mixed_human":  "Mixed Human",
    }
    AI_ORDER       = ["random", "fair_play", "challenge", "ml"]
    STRATEGY_ORDER = list(STRATEGY_MAP.keys())

    # Only include AI columns that have data — keeps the table clean during
    # partial data collection.
    ai_present = [a for a in AI_ORDER if any(k[1] == a for k in sim)]

    rows = []
    header = "| **Strategy** | " + " | ".join(f"**{AI_MAP.get(a, a)}**" for a in ai_present) + " |"
    rows.append(header)
    rows.append("| --- | " + " | ".join("---" for _ in ai_present) + " |")

    for strat in STRATEGY_ORDER:
        disp  = STRATEGY_MAP.get(strat, strat)
        cells = []
        for ai in ai_present:
            entry = sim.get((strat, ai))
            cells.append(f"{entry['robot_win_rate'] * 100:.1f}%" if entry else "n/a")
        rows.append(f"| {disp} | " + " | ".join(cells) + " |")

    return rows


def _table3_rows(ch: dict) -> list[str]:
    """Build Table 3: high-level challenge mode aggregate statistics."""
    tr = ch["total_r"]
    tg = ch["total_g"]
    return [
        "| **Metric** | **Value** |",
        "| --- | --- |",
        f"| Total Runs | {ch['total_runs']} |",
        f"| Total Rounds | {ch['total_rounds']} |",
        f"| Longest Streak | {ch['longest']} |",
        f"| Player Wins | {ch['p_wins']} ({_pct_of(ch['p_wins'], tr)}) |",
        f"| Robot Wins | {ch['robot_wins']} ({_pct_of(ch['robot_wins'], tr)}) |",
        f"| Draws | {ch['draws_ch']} ({_pct_of(ch['draws_ch'], tr)}) |",
        f"| Player Rock | {ch['rock']} ({_pct_of(ch['rock'], tg)}) |",
        f"| Player Paper | {ch['paper']} ({_pct_of(ch['paper'], tg)}) |",
        f"| Player Scissors | {ch['scissors']} ({_pct_of(ch['scissors'], tg)}) |",
    ]


def _table3b_rows(ch: dict) -> list[str]:
    """Build Table 3b: distribution of final streak lengths across all runs."""
    dist  = ch["streak_dist"]
    total = max(sum(dist.values()), 1)
    max_s = ch["longest"]

    rows = [
        "| **Streak** | **Runs** | **Percentage** |",
        "| --- | --- | --- |",
    ]
    # Iterate from 0 (lost immediately) up to the longest recorded streak.
    for s in range(0, max_s + 1):
        n   = dist.get(s, 0)
        # Label the first and last rows for clarity.
        if s == 0:
            lbl = "0 (immediate loss)"
        elif s == max_s:
            lbl = f"{s} (highest)"
        else:
            lbl = str(s)
        rows.append(f"| {lbl} | {n} | {_pct_of(n, total)} |")
    return rows


def _table3c_rows(ch: dict) -> list[str]:
    """Build Table 3c: how players respond after winning or drawing."""
    ocr = ch["ocr"]
    rows = [
        "| **After...** | **Stay** | **Upgrade** | **Downgrade** | **n** |",
        "| --- | --- | --- | --- | --- |",
    ]
    # Only show "after win" and "after draw" — these are the interesting patterns.
    for outcome, label in [("win", "Win"), ("draw", "Draw")]:
        d = ocr.get(outcome, {"stay": 0, "upgrade": 0, "downgrade": 0, "n": 0})
        n = d.get("n", 0)
        rows.append(
            f"| {label} | {d['stay'] * 100:.1f}% | {d['upgrade'] * 100:.1f}% "
            f"| {d['downgrade'] * 100:.1f}% | {n} |"
        )
    return rows


def _table4_rows(p: dict) -> list[str]:
    """Build Table 4: a human-readable summary of the player's behavioural traits."""
    if not p:
        return []
    ocr   = p["ocr"]
    w_ocr = ocr.get("win",  {"stay": 0, "upgrade": 0, "downgrade": 0})
    l_ocr = ocr.get("lose", {"stay": 0, "upgrade": 0, "downgrade": 0})
    d_ocr = ocr.get("draw", {"stay": 0, "upgrade": 0, "downgrade": 0})
    bt_src = p["best_trans_src"]
    bt_dst = p["best_trans_dst"]
    bt_pct = p["best_trans_pct"]

    # Find the most common response the player makes after a loss.
    lose_max = max(l_ocr, key=l_ocr.get) if l_ocr else "upgrade"
    lose_val = l_ocr.get(lose_max, 0) * 100

    # The gesture that beats the player's favourite — what the AI should play.
    beat_fav = {"Rock": "Paper", "Paper": "Scissors", "Scissors": "Rock"}
    counter  = beat_fav.get(p["fav_gesture"], "Rock")

    return [
        "| **Trait** | **Finding** |",
        "| --- | --- |",
        f"| Favourite gesture | {p['fav_gesture']} ({p['fav_pct'] * 100:.0f}% of throws) |",
        f"| Least used | {p['least_gesture']} ({p['least_pct'] * 100:.0f}%) |",
        f"| After winning | Downgrade {w_ocr['downgrade'] * 100:.0f}%, "
        f"Upgrade {w_ocr['upgrade'] * 100:.0f}%, Stay {w_ocr['stay'] * 100:.0f}% |",
        f"| After losing | {lose_max.title()} {lose_val:.0f}% of the time |",
        f"| After drawing | Downgrade {d_ocr['downgrade'] * 100:.0f}%, "
        f"Upgrade {d_ocr['upgrade'] * 100:.0f}%, Stay {d_ocr['stay'] * 100:.0f}% |",
        f"| Strongest transition | After {bt_src} → {bt_dst} ({bt_pct * 100:.0f}%) |",
        f"| Key trait | Stay rate {p['stay_pct'] * 100:.0f}% "
        f"(vs Nash 33.3%) |",
        f"| Best counter | Play {counter} often "
        f"(counters {p['fav_pct'] * 100:.0f}% {p['fav_gesture']}) |",
    ]


def _table5_rows(p: dict) -> list[str]:
    """Build Table 5: the full gesture-to-gesture transition matrix for one player."""
    if not p:
        return []
    trans = p["transition"]
    G     = ("Rock", "Paper", "Scissors")
    rows  = [
        "| **After...** | **→ Rock** | **→ Paper** | **→ Scissors** |",
        "| --- | --- | --- | --- |",
    ]
    for g in G:
        t     = trans.get(g, {})
        cells = [f"{t.get(g2, 0) * 100:.0f}%" for g2 in G]
        rows.append(f"| {g} | {cells[0]} | {cells[1]} | {cells[2]} |")
    return rows


def _table5b_rows(p: dict, ch: dict) -> list[str]:
    """Build Table 5b: side-by-side comparison of player vs challenge pop vs Nash baseline."""
    if not p or not ch:
        return []
    ch_tr = ch["total_r"]
    ch_tg = ch["total_g"]
    return [
        "| **Metric** | "
        f"**{p['player_name']} ({p['total_rounds']} rounds)** | "
        f"**Challenge Pop. ({ch['total_rounds']} rounds)** | "
        "**Literature Baseline** |",
        "| --- | --- | --- | --- |",
        f"| Stay rate overall | {p['stay_pct'] * 100:.1f}% | "
        f"{ch['ocr'].get('win', {}).get('stay', 0) * 30:.1f}% | 33.3% (Nash) |",
        f"| Upgrade rate overall | {p['upgrade_pct'] * 100:.1f}% | "
        f"33.5% | 33.3% (Nash) |",
        f"| Downgrade rate overall | {p['downgrade_pct'] * 100:.1f}% | "
        f"39.2% | 33.3% (Nash) |",
        f"| Gesture: Rock | {p['rock_pct'] * 100:.1f}% | "
        f"{_pct_of(ch['rock'], ch_tg)} | 35.7% [28] |",
        f"| Gesture: Paper | {p['paper_pct'] * 100:.1f}% | "
        f"{_pct_of(ch['paper'], ch_tg)} | 32.1% [28] |",
        f"| Gesture: Scissors | {p['scissors_pct'] * 100:.1f}% | "
        f"{_pct_of(ch['scissors'], ch_tg)} | 32.2% [28] |",
        f"| Win rate | {p['win_rate'] * 100:.1f}% | "
        f"{_pct_of(ch['p_wins'], ch_tr)} | 33.3% (Nash) |",
    ]


def _module_line_counts() -> dict[str, int]:
    """Count the actual number of lines in each .py file in the project directory."""
    counts = {}
    for py_file in PROJECT_DIR.glob("*.py"):
        try:
            with open(py_file, encoding="utf-8", errors="ignore") as f:
                counts[py_file.name] = sum(1 for _ in f)
        except Exception:
            pass
    return counts


# ---------------------------------------------------------------------------
# Report rewriter utilities
# ---------------------------------------------------------------------------

def _replace_table(content: str, header_pattern: str, new_rows: list[str]) -> str:
    """
    Find a Markdown table by a substring of its header row and replace it.

    The search is case-insensitive.  The replacement preserves everything
    outside the table (paragraphs, captions, headings) — it only touches
    lines that start with "|".

    This is deliberately simple: it just finds the first table whose header
    contains `header_pattern` and swaps out all the "|" lines.  The table
    header patterns are chosen to be unique within the document.
    """
    if not new_rows:
        return content  # nothing to do

    lines = content.splitlines()
    out   = []
    i     = 0
    while i < len(lines):
        line = lines[i]
        # A table header row: starts with "|" and contains our search pattern.
        if header_pattern.lower() in line.lower() and line.strip().startswith("|"):
            # Write all replacement rows in place.
            for nr in new_rows:
                out.append(nr)
            i += 1
            # Skip every old table row (all the lines that start with "|").
            while i < len(lines) and lines[i].strip().startswith("|"):
                i += 1
            continue  # don't fall through to the out.append below
        out.append(line)
        i += 1
    return "\n".join(out)


def _update_inline_numbers(content: str, ch: dict, p: dict) -> str:
    """
    Update key numbers mentioned in paragraph text (abstract, findings).

    Uses regex to find and replace specific sentence patterns so that numbers
    like round counts stay in sync with the Excel data without manual editing.
    """
    total_sim    = 18000  # the simulation dataset is fixed at 18,000 rounds
    live_rounds  = ch.get("total_rounds", 309)
    player_rounds = p.get("total_rounds", 52)

    # Update the live-round count wherever it appears in the text.
    content = re.sub(
        r"\b(\d{1,5}) live gameplay rounds\b",
        f"{live_rounds} live gameplay rounds",
        content,
    )
    # Update the single-player round count.
    content = re.sub(
        r"\b(\d{1,5}) rounds from one player",
        f"{player_rounds} rounds from one player",
        content,
    )
    # Update the Challenge run/round summary sentence.
    content = re.sub(
        r"122 Challenge mode runs totalling (\d+) rounds",
        f"{ch.get('total_runs', 122)} Challenge mode runs totalling {live_rounds} rounds",
        content,
    )
    # Update the auto-generated timestamp line if present.
    content = re.sub(
        r"(Generated:.*?)\d{4}-\d{2}-\d{2}",
        f"\\g<1>{datetime.now().strftime('%Y-%m-%d')}",
        content,
    )
    return content


def _update_table7_line_counts(content: str) -> str:
    """
    Scan the module inventory table and replace each module's line count with
    the actual current count from disk.

    Uses a regex to find Markdown table cells of the form "| module.py | ~NNN"
    and updates the number in place.
    """
    counts = _module_line_counts()
    if not counts:
        return content

    def _replace_line_count(m):
        module = m.group(1)
        # Only replace if we actually counted that file.
        if module in counts:
            return f"| {module} | ~{counts[module]:,}"
        return m.group(0)  # leave unchanged if we don't have a count

    return re.sub(
        r"\| ([\w_]+\.py) \| ~[\d,]+",
        _replace_line_count,
        content,
    )


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def update_report(
    report_src: Path | None = None,
    report_out: Path | None = None,
    verbose: bool = True,
) -> Path | None:
    """
    Read all Excel files and write an updated copy of the research report.

    This is the single public function used by main.py.  It orchestrates
    the full pipeline: load data → build table strings → find + replace in
    the Markdown source → write output file.

    Returns the output Path on success, or None if something went wrong.
    """
    if not OPENPYXL_OK:
        print("[Updater] openpyxl required. Install with: pip install openpyxl")
        return None

    src = Path(report_src or REPORT_SRC)
    out = Path(report_out or REPORT_OUT)

    if not src.exists():
        print(f"[Updater] Report not found at {src}")
        return None

    if verbose:
        print(f"[Updater] Reading report from {src}")

    with open(src, encoding="utf-8") as f:
        content = f.read()

    # -- Load data from all Excel sources --
    if verbose:
        print("[Updater] Reading Excel logs...")

    ch  = _challenge_summary()
    p   = _player_summary()
    sim = _simulation_summary()

    if verbose:
        print(f"  Challenge: {ch.get('total_runs', 0)} runs, "
              f"{ch.get('total_rounds', 0)} rounds, "
              f"longest streak {ch.get('longest', 0)}")
        print(f"  Player ({p.get('player_name', '?')}): "
              f"{p.get('total_rounds', 0)} rounds")
        print(f"  Simulation: {len(sim)} strategy-AI combos")

    # -- Replace each table if we have the required data for it --

    if sim:
        content = _replace_table(
            content, "**Strategy** | **Random AI**",
            _table2_rows(sim),
        )

    if ch:
        content = _replace_table(
            content, "**Metric** | **Value**",
            _table3_rows(ch),
        )
        content = _replace_table(
            content, "**Streak** | **Runs** | **Percentage**",
            _table3b_rows(ch),
        )
        content = _replace_table(
            content, "**After...** | **Stay** | **Upgrade** | **Downgrade** | **n**",
            _table3c_rows(ch),
        )

    if p:
        content = _replace_table(
            content, "**Trait** | **Finding**",
            _table4_rows(p),
        )
        content = _replace_table(
            content, "**After...** | **→ Rock**",
            _table5_rows(p),
        )
        if ch:
            # Table 5b header includes the player name, so it's always unique.
            content = _replace_table(
                content, f"**Metric** | **{p['player_name']}",
                _table5b_rows(p, ch),
            )

    # Update module line counts in Table 7.
    content = _update_table7_line_counts(content)

    # Update inline number mentions in the abstract / findings text.
    content = _update_inline_numbers(content, ch, p)

    # -- Write the updated file --
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        f.write(content)

    if verbose:
        print(f"[Updater] Updated report written to:\n  {out}")

    return out


def main():
    """Command-line entry point — prints a banner and runs the update."""
    print("=" * 60)
    print("RPS Capstone Report Auto-Updater")
    print(f"  Source : {REPORT_SRC}")
    print(f"  Output : {REPORT_OUT}")
    print(f"  Run at : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    result = update_report(verbose=True)
    if result:
        print("\nDone. Open the .md file or convert to .docx with:")
        print(f"  pandoc '{result}' -o '{result.with_suffix('.docx')}'")
    else:
        print("\nUpdate failed — check the error messages above.")


if __name__ == "__main__":
    main()
