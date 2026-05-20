# challenge_stats_logger.py
# --------------------------
# Saves stats for Challenge Mode to an Excel workbook on the Desktop.
# Tracks every round, every run (one full session), and lifetime totals.
# The workbook has three sheets:
#   - Summary          : lifetime totals (wins, streaks, gesture counts, etc.)
#   - Challenge_Runs   : one row per game session
#   - Challenge_Rounds : one row per individual round

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# Maps each Summary metric name to its row number in the Summary sheet (column B).
# This lets us look up a row by name instead of hardcoding row numbers everywhere.
SUMMARY_ROWS = {
    "longest_streak":       2,
    "total_runs":           3,
    "total_rounds":         4,
    "total_player_wins":    5,
    "total_draws":          6,
    "total_player_losses":  7,
    "workbook_path":        8,
    "last_updated":         9,
    # Gesture frequency counters (added v2)
    "player_rock_count":    10,
    "player_paper_count":   11,
    "player_scissors_count": 12,
    "robot_rock_count":     13,
    "robot_paper_count":    14,
    "robot_scissors_count": 15,
}

# Maps a gesture name to the corresponding Summary key for the player and robot.
# Used to increment the right counter without a chain of if/elif statements.
_PLAYER_GESTURE_KEY = {
    "Rock":     "player_rock_count",
    "Paper":    "player_paper_count",
    "Scissors": "player_scissors_count",
}
_ROBOT_GESTURE_KEY = {
    "Rock":     "robot_rock_count",
    "Paper":    "robot_paper_count",
    "Scissors": "robot_scissors_count",
}


class ChallengeStatsLogger:
    """
    Persistent logger for Challenge Mode.

    Creates and updates:
        ~/Desktop/CapStone/challenge_research_log.xlsx

    Sheets:
        Summary          -- lifetime totals including gesture frequencies
        Challenge_Runs   -- per-run summary with gesture counts
        Challenge_Rounds -- every individual round with full detail
    """

    def __init__(self, base_dir=None, workbook_name="challenge_research_log.xlsx"):
        # Default save location is ~/Desktop/CapStone/ — create it if missing.
        self.base_dir = Path(base_dir) if base_dir else Path.home() / "Desktop" / "CapStone"
        self.base_dir.mkdir(parents=True, exist_ok=True)

        self.workbook_path = self.base_dir / workbook_name

        # Context fields are written to every run and round row for filtering later.
        self.context = {
            "display_mode":      "Game",
            "camera_resolution": "640x480",
        }

        # Holds stats for the currently active run. None when no run is in progress.
        self.active_run = None

        # Create the workbook on first launch, then check for schema upgrades.
        self._ensure_workbook()
        self._migrate_workbook()

    # Allow the caller to update context (e.g. when the user changes camera mode).
    def update_context(self, display_mode=None, camera_resolution=None):
        if display_mode is not None:
            self.context["display_mode"] = display_mode
        if camera_resolution is not None:
            self.context["camera_resolution"] = camera_resolution

    def get_high_score(self):
        """Read and return the all-time longest streak from the Summary sheet."""
        try:
            wb = load_workbook(self.workbook_path)
            ws = wb["Summary"]
            value = ws[f"B{SUMMARY_ROWS['longest_streak']}"].value
            wb.close()
            return int(value or 0)
        except Exception:
            return 0

    def start_run(self):
        """
        Begin a new game run. If a run is already active, just return its ID.
        Increments the total_runs counter in Summary.
        Returns the run_id string.
        """
        # Don't start a second run if one is already going.
        if self.active_run is not None:
            return self.active_run["run_id"]

        timestamp = self._timestamp()
        run_id    = datetime.now().strftime("RUN-%Y%m%d-%H%M%S-%f")

        # Build the in-memory run record. Gesture counts track what was played this run.
        self.active_run = {
            "run_id":             run_id,
            "started_at":         timestamp,
            "rounds_played":      0,
            "wins":               0,
            "draws":              0,
            "losses":             0,
            "player_rocks":       0,
            "player_papers":      0,
            "player_scissors":    0,
            "robot_rocks":        0,
            "robot_papers":       0,
            "robot_scissors":     0,
            "reaction_times":     [],
            "display_mode":       self.context["display_mode"],
            "camera_resolution":  self.context["camera_resolution"],
        }

        # Write the incremented run count to the workbook immediately.
        wb = self._load_workbook()
        if wb is None:
            return run_id

        summary = wb["Summary"]
        self._increment_summary(summary, "total_runs", 1)
        self._set_summary(summary, "last_updated", timestamp)
        self._safe_save_and_close(wb)

        return run_id

    def log_round(
        self,
        round_number,
        player_gesture,
        robot_gesture,
        round_result,
        streak_after_round,
        high_score_after_round,
        ai_predicted_move=None,
        ai_effective_skill=None,
        reaction_time_ms=None,
        previous_player_gesture=None,
        player_response_type=None,
        emotion=None,
        emotion_confidence=None,
        smile_score=None,
        surprise_score=None,
        frustration_score=None
    ):
        """
        Record a single round to the Challenge_Rounds sheet and update Summary totals.
        Automatically starts a run if none is active.
        """
        run_id    = self.start_run()
        timestamp = self._timestamp()

        wb = self._load_workbook()
        if wb is None:
            return

        # Append the full round detail to the rounds sheet.
        rounds_ws = wb["Challenge_Rounds"]
        rounds_ws.append([
            timestamp,
            run_id,
            round_number,
            player_gesture,
            robot_gesture,
            round_result,
            streak_after_round,
            high_score_after_round,
            self.context["camera_resolution"],
            self.context["display_mode"],
            ai_predicted_move,
            ai_effective_skill,
            reaction_time_ms,
            previous_player_gesture,
            player_response_type,
            emotion,
            emotion_confidence,
            smile_score,
            surprise_score,
            frustration_score,
        ])

        summary = wb["Summary"]
        self._increment_summary(summary, "total_rounds", 1)

        # Increment the appropriate win/draw/loss counter for the round outcome.
        if round_result == "player_win":
            self._increment_summary(summary, "total_player_wins", 1)
            if self.active_run is not None:
                self.active_run["wins"] += 1
        elif round_result == "draw":
            self._increment_summary(summary, "total_draws", 1)
            if self.active_run is not None:
                self.active_run["draws"] += 1
        elif round_result == "robot_win":
            self._increment_summary(summary, "total_player_losses", 1)
            if self.active_run is not None:
                self.active_run["losses"] += 1

        # Increment the global gesture frequency counters using the lookup tables.
        player_key = _PLAYER_GESTURE_KEY.get(player_gesture)
        if player_key:
            self._increment_summary(summary, player_key, 1)

        robot_key = _ROBOT_GESTURE_KEY.get(robot_gesture)
        if robot_key:
            self._increment_summary(summary, robot_key, 1)

        # Also track per-run gesture counts in memory.
        if self.active_run is not None:
            run = self.active_run
            if player_gesture == "Rock":
                run["player_rocks"] += 1
            elif player_gesture == "Paper":
                run["player_papers"] += 1
            elif player_gesture == "Scissors":
                run["player_scissors"] += 1

            if robot_gesture == "Rock":
                run["robot_rocks"] += 1
            elif robot_gesture == "Paper":
                run["robot_papers"] += 1
            elif robot_gesture == "Scissors":
                run["robot_scissors"] += 1

        # Update the all-time high score if this round beat it.
        current_best = self._get_summary_value(summary, "longest_streak")
        if high_score_after_round > current_best:
            self._set_summary(summary, "longest_streak", high_score_after_round)

        self._set_summary(summary, "last_updated", timestamp)

        # Update in-memory run counters.
        if self.active_run is not None:
            self.active_run["rounds_played"] += 1
            if reaction_time_ms is not None:
                self.active_run["reaction_times"].append(reaction_time_ms)

        self._safe_save_and_close(wb)

    def finalize_run(self, final_streak, status="completed"):
        """
        Called when a game session ends. Writes a summary row to Challenge_Runs
        and updates the high score if needed. Clears the active run.
        """
        # Grab and clear the active run immediately so start_run can begin a new one.
        run = self.active_run
        self.active_run = None
        if run is None:
            return

        timestamp = self._timestamp()

        wb = self._load_workbook()
        if wb is None:
            return

        # Average reaction time across all rounds in this run (None if no data).
        reaction_times = run["reaction_times"]
        avg_reaction_ms = (
            round(sum(reaction_times) / len(reaction_times), 1)
            if reaction_times else None
        )

        # Write the per-run summary row.
        runs_ws = wb["Challenge_Runs"]
        runs_ws.append([
            run["run_id"],
            run["started_at"],
            timestamp,
            status,
            final_streak,
            run["rounds_played"],
            run["wins"],
            run["draws"],
            run["losses"],
            run["player_rocks"],
            run["player_papers"],
            run["player_scissors"],
            run["robot_rocks"],
            run["robot_papers"],
            run["robot_scissors"],
            avg_reaction_ms,
            run["camera_resolution"],
            run["display_mode"],
        ])

        # Update the all-time high score once more using the final streak.
        summary = wb["Summary"]
        current_best = self._get_summary_value(summary, "longest_streak")
        if final_streak > current_best:
            self._set_summary(summary, "longest_streak", final_streak)

        self._set_summary(summary, "last_updated", timestamp)
        self._safe_save_and_close(wb)

    # ------------------------------------------------------------------
    # Workbook creation (runs on first launch if file doesn't exist)
    # ------------------------------------------------------------------

    def _ensure_workbook(self):
        """
        Create the Excel workbook with all three sheets if it doesn't exist yet.
        Sets up headers, default values, column widths, and frozen header rows.
        """
        if self.workbook_path.exists():
            return

        wb = Workbook()

        # ── Summary sheet ──────────────────────────────────────────────
        summary_ws = wb.active
        summary_ws.title = "Summary"
        summary_ws["A1"] = "Metric"
        summary_ws["B1"] = "Value"

        # Write each metric label and its starting value of 0.
        summary_ws["A2"]  = "longest_streak";       summary_ws["B2"]  = 0
        summary_ws["A3"]  = "total_runs";            summary_ws["B3"]  = 0
        summary_ws["A4"]  = "total_rounds";          summary_ws["B4"]  = 0
        summary_ws["A5"]  = "total_player_wins";     summary_ws["B5"]  = 0
        summary_ws["A6"]  = "total_draws";           summary_ws["B6"]  = 0
        summary_ws["A7"]  = "total_player_losses";   summary_ws["B7"]  = 0
        summary_ws["A8"]  = "workbook_path";         summary_ws["B8"]  = str(self.workbook_path)
        summary_ws["A9"]  = "last_updated";          summary_ws["B9"]  = self._timestamp()
        summary_ws["A10"] = "player_rock_count";     summary_ws["B10"] = 0
        summary_ws["A11"] = "player_paper_count";    summary_ws["B11"] = 0
        summary_ws["A12"] = "player_scissors_count"; summary_ws["B12"] = 0
        summary_ws["A13"] = "robot_rock_count";      summary_ws["B13"] = 0
        summary_ws["A14"] = "robot_paper_count";     summary_ws["B14"] = 0
        summary_ws["A15"] = "robot_scissors_count";  summary_ws["B15"] = 0

        # ── Challenge_Runs sheet: one row per game session ─────────────
        runs_ws = wb.create_sheet("Challenge_Runs")
        runs_ws.append([
            "run_id", "started_at", "ended_at", "status", "final_streak",
            "rounds_played", "wins", "draws", "losses",
            "player_rocks", "player_papers", "player_scissors",
            "robot_rocks",  "robot_papers",  "robot_scissors",
            "avg_reaction_ms", "camera_resolution", "display_mode",
        ])

        # ── Challenge_Rounds sheet: one row per individual round ───────
        rounds_ws = wb.create_sheet("Challenge_Rounds")
        rounds_ws.append([
            "timestamp", "run_id", "round_number",
            "player_gesture", "robot_gesture", "round_result",
            "streak_after_round", "high_score_after_round",
            "camera_resolution", "display_mode",
            "ai_predicted_move", "ai_effective_skill", "reaction_time_ms",
            "previous_player_gesture", "player_response_type",
            "emotion", "emotion_confidence",
            "smile_score", "surprise_score", "frustration_score",
        ])

        # Apply header styling and freeze the top row on all sheets.
        self._format_sheet(summary_ws)
        self._format_sheet(runs_ws)
        self._format_sheet(rounds_ws)

        # Set column widths so data is readable without manual resizing.
        self._set_column_widths(summary_ws, {"A": 24, "B": 44})
        self._set_column_widths(runs_ws, {
            "A": 28, "B": 22, "C": 22, "D": 14, "E": 14,
            "F": 14, "G": 10, "H": 10, "I": 10, "J": 14,
            "K": 14, "L": 16, "M": 14, "N": 14, "O": 16,
            "P": 18, "Q": 18, "R": 14,
        })
        self._set_column_widths(rounds_ws, {
            "A": 22, "B": 28, "C": 14, "D": 16, "E": 16,
            "F": 16, "G": 18, "H": 20, "I": 18, "J": 14,
            "K": 18, "L": 18, "M": 18, "N": 22, "O": 20,
            "P": 14, "Q": 18, "R": 14, "S": 14, "T": 16,
        })

        self._safe_save_and_close(wb)

    # ------------------------------------------------------------------
    # Migration: upgrades existing workbooks without losing data
    # ------------------------------------------------------------------

    def _migrate_workbook(self):
        """
        Called every startup. Checks whether the on-disk workbook is missing any
        columns or rows that were added in later versions, and adds them if so.
        Safe to run on an already up-to-date workbook — it checks before writing.
        """
        if not self.workbook_path.exists():
            return

        wb = self._load_workbook()
        if wb is None:
            return

        changed = False
        summary = wb["Summary"]

        # ── Summary: add gesture frequency rows if not present ─────────
        # Row A10 should be "player_rock_count"; if it's something else this
        # workbook was created before v2 and needs the new rows added.
        if summary["A10"].value != "player_rock_count":
            new_rows = {
                10: "player_rock_count",
                11: "player_paper_count",
                12: "player_scissors_count",
                13: "robot_rock_count",
                14: "robot_paper_count",
                15: "robot_scissors_count",
            }
            for row_num, label in new_rows.items():
                summary[f"A{row_num}"] = label
                # Only zero-fill if the cell is empty — don't overwrite real data.
                if summary[f"B{row_num}"].value is None:
                    summary[f"B{row_num}"] = 0
            changed = True
            print("[ChallengeStats] Migrated Summary: added gesture frequency rows.")

        # ── Challenge_Runs: update header if columns were added later ──
        runs_ws     = wb["Challenge_Runs"]
        runs_header = [cell.value for cell in runs_ws[1]]

        target_runs_header = [
            "run_id", "started_at", "ended_at", "status", "final_streak",
            "rounds_played", "wins", "draws", "losses",
            "player_rocks", "player_papers", "player_scissors",
            "robot_rocks",  "robot_papers",  "robot_scissors",
            "avg_reaction_ms", "camera_resolution", "display_mode",
        ]

        if runs_header != target_runs_header:
            # Overwrite the header row with the full target list.
            for col_idx, value in enumerate(target_runs_header, start=1):
                runs_ws.cell(row=1, column=col_idx, value=value)

            # Re-apply the header styling so the new cells match the old ones.
            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in runs_ws[1]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = Alignment(horizontal="center")

            self._set_column_widths(runs_ws, {
                "J": 14, "K": 14, "L": 16,
                "M": 14, "N": 14, "O": 16,
                "P": 18, "Q": 18, "R": 14,
            })
            changed = True
            print("[ChallengeStats] Migrated Challenge_Runs: updated header columns.")

        # ── Challenge_Rounds: update header if emotion columns were added ──
        rounds_ws     = wb["Challenge_Rounds"]
        rounds_header = [cell.value for cell in rounds_ws[1]]

        target_rounds_header = [
            "timestamp", "run_id", "round_number",
            "player_gesture", "robot_gesture", "round_result",
            "streak_after_round", "high_score_after_round",
            "camera_resolution", "display_mode",
            "ai_predicted_move", "ai_effective_skill", "reaction_time_ms",
            "previous_player_gesture", "player_response_type",
            "emotion", "emotion_confidence",
            "smile_score", "surprise_score", "frustration_score",
        ]

        if rounds_header != target_rounds_header:
            for col_idx, value in enumerate(target_rounds_header, start=1):
                rounds_ws.cell(row=1, column=col_idx, value=value)

            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in rounds_ws[1]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = Alignment(horizontal="center")

            self._set_column_widths(rounds_ws, {
                "K": 18, "L": 18, "M": 18,
                "N": 22, "O": 20,
                "P": 14, "Q": 18, "R": 14, "S": 14, "T": 16,
            })
            changed = True
            print("[ChallengeStats] Migrated Challenge_Rounds: added emotion columns.")

        # Only save if we actually changed something.
        if changed:
            self._safe_save_and_close(wb)
        else:
            wb.close()

    # ------------------------------------------------------------------
    # Workbook helpers
    # ------------------------------------------------------------------

    def _load_workbook(self):
        """Open the workbook from disk. Returns None if the file is locked or corrupt."""
        try:
            return load_workbook(self.workbook_path)
        except PermissionError:
            print(
                "[ChallengeStats] Could not open challenge_research_log.xlsx. "
                "Please close the workbook in Excel and try again."
            )
            return None
        except Exception as exc:
            print(f"[ChallengeStats] Failed to open workbook: {exc}")
            return None

    def _safe_save_and_close(self, wb):
        """Save the workbook and close it. Prints a friendly message if Excel has it locked."""
        try:
            wb.save(self.workbook_path)
        except PermissionError:
            print(
                "[ChallengeStats] Could not save challenge_research_log.xlsx. "
                "Please close the workbook in Excel."
            )
        except Exception as exc:
            print(f"[ChallengeStats] Failed to save workbook: {exc}")
        finally:
            wb.close()

    def _increment_summary(self, summary_ws, key, amount):
        """Add `amount` to the numeric cell for the given Summary metric."""
        row     = SUMMARY_ROWS[key]
        current = summary_ws[f"B{row}"].value or 0
        summary_ws[f"B{row}"] = int(current) + amount

    def _set_summary(self, summary_ws, key, value):
        """Overwrite the value cell for the given Summary metric."""
        row = SUMMARY_ROWS[key]
        summary_ws[f"B{row}"] = value

    def _get_summary_value(self, summary_ws, key):
        """Read the integer value of a Summary metric (returns 0 if empty)."""
        row   = SUMMARY_ROWS[key]
        value = summary_ws[f"B{row}"].value
        return int(value or 0)

    def _format_sheet(self, ws):
        """
        Style the header row with a dark blue background and bold white text,
        then freeze it so it stays visible while scrolling.
        """
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    def _set_column_widths(self, ws, widths):
        """Set column widths from a dict mapping column letter -> width in characters."""
        for column_letter, width in widths.items():
            ws.column_dimensions[column_letter].width = width

    def _timestamp(self):
        """Return the current date and time as a readable string."""
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
