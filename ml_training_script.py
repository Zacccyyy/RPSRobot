"""
ml_training_script.py
=====================
Standalone script that trains the ML player-prediction model from real
gameplay data (and optionally simulation data) stored in Excel workbooks.

WHERE IT FITS:
    challenge_research_log.xlsx  (written by challenge_mode_state.py)
    simulation_results.xlsx      (written by simulation_mode.py)
            |
    ml_training_script.py  (this file — reads, trains, evaluates, saves)
            |
    rps_ml_model.pkl  (loaded at runtime by MLPredictionAI)

HOW TO RUN:
    python ml_training_script.py

The CONFIG SECTION below lets you change the model type, lookback window,
train/test split, and whether to include simulation data.

REQUIREMENTS:
    pip install scikit-learn numpy openpyxl
    (already in requirements.txt)

OUTPUT:
    ~/Desktop/CapStone/rps_ml_model.pkl  — trained model file
    Terminal printout with accuracy and top feature importances
"""

import sys
from pathlib import Path
from collections import defaultdict, Counter

from openpyxl import load_workbook

from ml_feature_extractor import build_training_set, get_feature_names, GESTURE_INDEX
from ml_model import RPSModel, INDEX_TO_GESTURE


# ===========================================================================
# CONFIG — edit these values to tune training behaviour
# ===========================================================================

WORKBOOK_PATH     = Path.home() / "Desktop" / "CapStone" / "challenge_research_log.xlsx"
SIMULATION_PATH   = Path.home() / "Desktop" / "CapStone" / "simulation_results.xlsx"
MODEL_OUTPUT_PATH = Path.home() / "Desktop" / "CapStone" / "rps_ml_model.pkl"

MODEL_TYPE          = "logistic"   # "logistic" (safer, fast) or "forest" (may overfit)
LOOKBACK            = 3            # how many past rounds to encode as features
TEST_SPLIT          = 0.20         # fraction of data held out for evaluation
MIN_SAMPLES         = 20           # abort early if we have fewer samples than this
USE_SIMULATION_DATA = True         # merge simulation data with real gameplay


# ===========================================================================
# Data loading helpers
# ===========================================================================

def load_rounds_from_excel(workbook_path):
    """
    Read the 'Challenge_Rounds' sheet from the research log workbook.

    Expected columns: run_id, round_number, player_gesture, robot_gesture,
    round_result. The 'round_result' values ("player_win", "robot_win", "draw")
    are translated to internal format ("win", "lose", "draw").

    Returns:
        dict of {run_id: [list of round dicts]}, sorted by round_number.
        Returns an empty dict if the file is missing or malformed.
    """
    if not workbook_path.exists():
        print(f"[Training] Workbook not found: {workbook_path}")
        return {}

    try:
        wb = load_workbook(workbook_path, read_only=True)
    except Exception as exc:
        print(f"[Training] Could not open workbook: {exc}")
        return {}

    if "Challenge_Rounds" not in wb.sheetnames:
        print("[Training] Challenge_Rounds sheet not found.")
        wb.close()
        return {}

    ws = wb["Challenge_Rounds"]

    # Parse the header row into a column-name -> index mapping
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col    = {name: i for i, name in enumerate(header)}

    # Verify all required columns exist before processing any rows
    required = ["run_id", "round_number", "player_gesture", "robot_gesture", "round_result"]
    for field in required:
        if field not in col:
            print(f"[Training] Missing column: {field}")
            wb.close()
            return {}

    # Translate the workbook's result strings to what the feature extractor expects
    result_to_outcome = {
        "player_win": "win",
        "robot_win":  "lose",
        "draw":       "draw",
    }

    rounds_by_run = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        run_id         = row[col["run_id"]]
        player_gesture = row[col["player_gesture"]]
        round_result   = row[col["round_result"]]

        # Skip rows with unrecognised gestures (e.g. data entry errors)
        if player_gesture not in GESTURE_INDEX:
            continue

        player_outcome = result_to_outcome.get(round_result)
        if player_outcome is None:
            continue   # skip rows with unrecognised result strings

        record = {
            "round_number":   row[col["round_number"]],
            "player_gesture": player_gesture,
            "robot_gesture":  row[col["robot_gesture"]],
            "player_outcome": player_outcome,
        }

        # Reaction time is optional — older data may not have this column
        if "reaction_time_ms" in col:
            record["reaction_time_ms"] = row[col["reaction_time_ms"]]

        rounds_by_run[run_id].append(record)

    wb.close()

    # Sort each run chronologically so the lookback window is in the right order
    for run_id in rounds_by_run:
        rounds_by_run[run_id].sort(key=lambda r: r["round_number"])

    return dict(rounds_by_run)


def load_rounds_from_simulation(simulation_path):
    """
    Read the 'Sim_Rounds' sheet from the simulation results workbook.

    Simulation data uses the same round dict format as real gameplay, but
    run IDs are prefixed with "SIM_" so they never collide when merged.

    Returns:
        dict of {run_id: [list of round dicts]}, or an empty dict if missing.
    """
    if not simulation_path.exists():
        print(f"[Training] Simulation file not found: {simulation_path}")
        return {}

    try:
        wb = load_workbook(simulation_path, read_only=True)
    except Exception as exc:
        print(f"[Training] Could not open simulation file: {exc}")
        return {}

    if "Sim_Rounds" not in wb.sheetnames:
        print("[Training] Sim_Rounds sheet not found.")
        wb.close()
        return {}

    ws = wb["Sim_Rounds"]

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col    = {name: i for i, name in enumerate(header)}

    # Simulation data already uses "win" / "lose" / "draw" directly
    required = ["run_id", "round_number", "player_gesture", "robot_gesture", "player_outcome"]
    for field in required:
        if field not in col:
            print(f"[Training] Sim_Rounds missing column: {field}")
            wb.close()
            return {}

    rounds_by_run = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        run_id         = row[col["run_id"]]
        player_gesture = row[col["player_gesture"]]
        player_outcome = row[col["player_outcome"]]

        # Skip rows with unrecognised values
        if player_gesture not in GESTURE_INDEX:
            continue
        if player_outcome not in ("win", "lose", "draw"):
            continue

        record = {
            "round_number":   row[col["round_number"]],
            "player_gesture": player_gesture,
            "robot_gesture":  row[col["robot_gesture"]],
            "player_outcome": player_outcome,
        }

        # Prefix run IDs so they don't collide with real-gameplay run IDs
        rounds_by_run[f"SIM_{run_id}"].append(record)

    wb.close()

    # Sort each run chronologically
    for run_id in rounds_by_run:
        rounds_by_run[run_id].sort(key=lambda r: r["round_number"])

    return dict(rounds_by_run)


# ===========================================================================
# Main training pipeline
# ===========================================================================

def main():
    print("=" * 60)
    print("RPS ML Training Script")
    print("=" * 60)
    print()

    # --- Load real gameplay data ---
    print(f"Loading gameplay data from: {WORKBOOK_PATH}")
    rounds_by_run = load_rounds_from_excel(WORKBOOK_PATH)

    real_runs   = len(rounds_by_run)
    real_rounds = sum(len(r) for r in rounds_by_run.values())
    print(f"  Real gameplay: {real_runs} runs, {real_rounds} rounds.")

    # --- Optionally merge simulation data ---
    if USE_SIMULATION_DATA:
        print(f"Loading simulation data from: {SIMULATION_PATH}")
        sim_rounds = load_rounds_from_simulation(SIMULATION_PATH)

        sim_run_count   = len(sim_rounds)
        sim_round_count = sum(len(r) for r in sim_rounds.values())
        print(f"  Simulation: {sim_run_count} runs, {sim_round_count} rounds.")

        # Merge simulation runs into the main dict (SIM_ prefix prevents collisions)
        rounds_by_run.update(sim_rounds)
    else:
        print("Simulation data: skipped (USE_SIMULATION_DATA = False)")

    total_runs   = len(rounds_by_run)
    total_rounds = sum(len(r) for r in rounds_by_run.values())
    print(f"  Combined: {total_runs} runs, {total_rounds} total rounds.")
    print()

    # Bail out early if we don't have enough data to train anything meaningful
    if total_rounds < MIN_SAMPLES:
        print(
            f"Not enough data to train. Need at least {MIN_SAMPLES} rounds, "
            f"have {total_rounds}."
        )
        print("Play more Challenge mode games to collect training data!")
        sys.exit(0)

    # --- Build feature matrix -----------------------------------------------
    print(f"Building features (lookback={LOOKBACK})...")
    X, y = build_training_set(rounds_by_run, lookback=LOOKBACK)

    print(f"Training samples: {len(X)}")
    print(f"Feature vector size: {len(X[0]) if X else 0}")

    # Show class distribution so we can spot badly imbalanced data
    dist = Counter(y)
    for cls_idx in sorted(dist):
        gesture = INDEX_TO_GESTURE.get(cls_idx, str(cls_idx))
        print(f"  {gesture}: {dist[cls_idx]} ({dist[cls_idx] / len(y) * 100:.1f}%)")
    print()

    # --- Train / test split --------------------------------------------------
    from sklearn.model_selection import train_test_split

    if len(X) < 10:
        # Too few samples to hold any out safely — train and evaluate on all data
        print("Very few samples — training on all data (no test split).")
        X_train, y_train = X, y
        X_test,  y_test  = X, y
    else:
        X_train, X_test, y_train, y_test = train_test_split(
            X, y,
            test_size=TEST_SPLIT,
            random_state=42,
            # stratify ensures each class appears proportionally in both splits
            stratify=y if len(set(y)) > 1 else None,
        )

    print(f"Train: {len(X_train)} samples")
    print(f"Test:  {len(X_test)} samples")
    print()

    # --- Train ---------------------------------------------------------------
    print(f"Training model (type={MODEL_TYPE})...")
    model = RPSModel(lookback=LOOKBACK)
    model.train(X_train, y_train, model_type=MODEL_TYPE)
    print()

    # --- Evaluate on held-out test set --------------------------------------
    print("Evaluation on test set:")
    results = model.evaluate(X_test, y_test)
    print(f"  Accuracy: {results['accuracy']:.1%}")
    print(f"  Samples:  {results['samples']}")
    print()

    # Per-class breakdown:
    #   precision = "when we predicted X, were we right?"
    #   recall    = "of all the actual X throws, did we catch them?"
    report = results.get("report", {})
    for gesture in GESTURE_INDEX:
        if gesture in report:
            r = report[gesture]
            print(
                f"  {gesture:10s}  "
                f"precision={r['precision']:.2f}  "
                f"recall={r['recall']:.2f}  "
                f"f1={r['f1-score']:.2f}  "
                f"n={int(r['support'])}"
            )
    print()

    # --- How much better are we than random? ---------------------------------
    # A random bot wins 1/3 of the time; lift shows how much the model adds
    random_accuracy = 1.0 / 3.0
    lift            = results["accuracy"] - random_accuracy
    print(f"Random baseline:  {random_accuracy:.1%}")
    print(f"Model accuracy:   {results['accuracy']:.1%}")
    print(f"Lift over random: {lift:+.1%}")
    print()

    # --- Top feature importances ---------------------------------------------
    importance = model.get_feature_importance()
    if importance:
        print("Top 10 features by importance:")
        for name, score in importance[:10]:
            print(f"  {name:30s}  {score:.4f}")
        print()

    # --- Save ----------------------------------------------------------------
    print(f"Saving model to: {MODEL_OUTPUT_PATH}")
    model.save(str(MODEL_OUTPUT_PATH))
    print()

    print("Done! You can now use this model in the game by passing")
    print(f'  model_path="{MODEL_OUTPUT_PATH}"')
    print("to MLPredictionAI.")


if __name__ == "__main__":
    main()
