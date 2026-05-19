"""
research_report.py
==================
Generates the AI comparison section of the RPS capstone research report.

This script sits at the top of the research pipeline.  It runs the simulation
engine (simulation_mode.py) for each AI type, collects the win/loss statistics,
and writes a formatted multi-sheet Excel report that can be opened in Excel or
Google Sheets.

Three (optionally four) AI types are compared:
    1. Random baseline  — expected ~33% robot win rate, used as a sanity check
    2. Heuristic FairPlay — pattern-exploitation AI using heuristic rules
    3. Heuristic Challenge — harder version of the heuristic AI
    4. ML Prediction     — Random-Forest model (only run if the .pkl file exists)

Each AI is tested against every player strategy defined in simulation_mode.py.
RUNS_PER_COMBO independent runs are averaged to smooth out variance.

Usage:
    python research_report.py

Output:
    ~/Desktop/CapStone/research_comparison_report.xlsx
"""

import os
import sys
import time
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from simulation_mode import run_simulation, PLAYER_STRATEGIES

# Path to the trained ML model (may not exist — that is handled gracefully).
ML_MODEL_PATH = os.path.join(
    os.path.expanduser("~"), "Desktop", "CapStone", "rps_ml_model.pkl"
)
# Where the finished Excel report is saved.
OUTPUT_PATH = Path.home() / "Desktop" / "CapStone" / "research_comparison_report.xlsx"

# How many independent simulation runs to average per strategy+AI combination.
# More runs = smoother numbers but slower to generate.
RUNS_PER_COMBO  = 10
ROUNDS_PER_RUN  = 100


# ===========================================================================
# Excel cell formatting constants
# ===========================================================================
# These are defined once here and reused throughout the sheet builders below
# so we never have to hardcode colours or font sizes in multiple places.

HEADER_FILL    = PatternFill("solid", fgColor="1F4E78")   # dark navy
HEADER_FONT    = Font(color="FFFFFF", bold=True, size=11)  # white bold
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")   # light blue
SUBHEADER_FONT = Font(bold=True, size=11)
TITLE_FONT     = Font(bold=True, size=14)
SUBTITLE_FONT  = Font(bold=True, size=12, color="1F4E78")
BODY_FONT      = Font(size=11)
GOOD_FILL      = PatternFill("solid", fgColor="C6EFCE")   # green  — AI winning well
BAD_FILL       = PatternFill("solid", fgColor="FFC7CE")   # red    — AI barely beating random
NEUTRAL_FILL   = PatternFill("solid", fgColor="FFEB9C")   # yellow — mid-range

# A thin border applied to every data cell for readability.
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _format_header_row(ws, row_num):
    """Apply the dark navy header style to every cell in a given row."""
    for cell in ws[row_num]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER


def _format_data_cell(ws, row, col, value, fmt=None):
    """
    Write `value` to a cell and apply the standard body style.

    `fmt` controls the number format:
      "pct"  → percentage with one decimal  (0.0%)
      "dec1" → number with one decimal      (0.0)
      None   → no number format (general)
    """
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = BODY_FONT
    cell.alignment = Alignment(horizontal="center")
    cell.border    = THIN_BORDER
    if fmt == "pct":
        cell.number_format = "0.0%"
    elif fmt == "dec1":
        cell.number_format = "0.0"
    return cell


def _auto_width(ws, min_width=10, max_width=22):
    """
    Set each column's width to fit its longest content, within [min, max].

    openpyxl doesn't auto-size columns, so we iterate all cells and measure
    the string length ourselves.
    """
    for col in ws.columns:
        letter  = get_column_letter(col[0].column)
        max_len = min_width
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(max_len, max_width)


# ===========================================================================
# Simulation runner
# ===========================================================================

def run_all_comparisons():
    """
    Run the full simulation suite for all AI types and collect the results.

    Returns:
      all_results  : dict mapping AI label → simulation result dict
      ml_available : bool — True if the ML model was found and run
    """
    # Base list of AI configurations to test.
    # Each tuple is (display label, list of AI keys passed to run_simulation).
    ai_configs = [
        ("Random",               ["random"]),
        ("Heuristic (FairPlay)", ["fair_play"]),
        ("Heuristic (Challenge)", ["challenge"]),
    ]

    # The ML model is optional — only add it if the .pkl file exists on disk.
    ml_available = os.path.exists(ML_MODEL_PATH)
    if ml_available:
        ai_configs.append(("ML Prediction", ["ml"]))
        print(f"ML model found at {ML_MODEL_PATH}")
    else:
        print(f"ML model NOT found at {ML_MODEL_PATH} — skipping ML comparison.")

    all_results = {}

    # Run the simulation for each AI type in turn.
    for ai_label, ai_list in ai_configs:
        print(f"\nRunning: {ai_label}...")
        results = run_simulation(
            player_strategies=PLAYER_STRATEGIES,
            ai_opponents=ai_list,
            runs_per_combo=RUNS_PER_COMBO,
            rounds_per_run=ROUNDS_PER_RUN,
            save_excel=False,  # we build our own report here
        )
        all_results[ai_label] = results

    return all_results, ml_available


# ===========================================================================
# Excel report builder
# ===========================================================================

def build_report(all_results, ml_available, output_path):
    """
    Create a new Excel workbook and populate it with all four (or five) sheets.

    Sheet order:
      1. Overview          — one row per AI, averaged across all strategies
      2. AI Comparison     — robot win rate grid: strategy x AI
      3. Per Strategy Detail — flat table of every individual combo
      4. Key Findings      — auto-generated text summary
      5. ML Model Details  — (only if ml_available) feature importances etc.
    """
    wb = Workbook()

    _build_overview_sheet(wb, all_results)
    _build_comparison_table(wb, all_results)
    _build_per_strategy_sheet(wb, all_results)
    _build_key_findings_sheet(wb, all_results, ml_available)

    if ml_available:
        _build_ml_details_sheet(wb)

    # Save the workbook.  PermissionError usually means the file is open in Excel.
    try:
        wb.save(output_path)
        print(f"\nReport saved to: {output_path}")
    except PermissionError:
        print(f"\nCould not save — close {output_path} in Excel and retry.")
    finally:
        wb.close()


def _build_overview_sheet(wb, all_results):
    """
    Populate the 'Overview' sheet with a summary row for each AI type.

    Columns: AI type | avg robot WR | avg player WR | avg draw rate | total rounds.
    Robot win rates above 38% are highlighted green; below 30% are red.
    """
    ws = wb.active
    ws.title = "Overview"

    # Title and metadata block at the top.
    ws.merge_cells("A1:F1")
    ws["A1"]       = "RPS AI Research Comparison Report"
    ws["A1"].font  = TITLE_FONT

    ws["A3"] = "Generated:"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Runs per combination:"
    ws["B4"] = RUNS_PER_COMBO
    ws["A5"] = "Rounds per run:"
    ws["B5"] = ROUNDS_PER_RUN
    ws["A6"] = "Player strategies tested:"
    ws["B6"] = len(PLAYER_STRATEGIES)

    # Header row for the data grid.
    row = 8
    ws.cell(row=row, column=1, value="AI Type").font           = SUBHEADER_FONT
    ws.cell(row=row, column=2, value="Avg Robot Win Rate").font = SUBHEADER_FONT
    ws.cell(row=row, column=3, value="Avg Player Win Rate").font = SUBHEADER_FONT
    ws.cell(row=row, column=4, value="Avg Draw Rate").font      = SUBHEADER_FONT
    ws.cell(row=row, column=5, value="Total Rounds").font       = SUBHEADER_FONT
    _format_header_row(ws, row)

    row += 1
    for ai_label, results in all_results.items():
        combos = results["combo_results"]
        if not combos:
            continue  # skip if this AI produced no data

        # Average the per-combo rates across all player strategies.
        avg_rwr = sum(c["robot_win_rate"]  for c in combos) / len(combos)
        avg_pwr = sum(c["player_win_rate"] for c in combos) / len(combos)
        avg_dr  = sum(c["draw_rate"]       for c in combos) / len(combos)
        total_r = results["total_rounds"]

        _format_data_cell(ws, row, 1, ai_label)
        _format_data_cell(ws, row, 2, avg_rwr, "pct")
        _format_data_cell(ws, row, 3, avg_pwr, "pct")
        _format_data_cell(ws, row, 4, avg_dr,  "pct")
        _format_data_cell(ws, row, 5, total_r)

        # Colour-code the robot win rate cell to make good/bad results stand out.
        cell = ws.cell(row=row, column=2)
        if avg_rwr > 0.38:
            cell.fill = GOOD_FILL    # clearly beating random — green
        elif avg_rwr < 0.30:
            cell.fill = BAD_FILL     # barely above random or worse — red

        row += 1

    _auto_width(ws)


def _build_comparison_table(wb, all_results):
    """
    Populate the 'AI Comparison' sheet: a grid of robot win rates.

    Rows = player strategies.  Column pairs = (robot WR, avg max streak) per AI.
    """
    ws        = wb.create_sheet("AI Comparison")
    ai_labels = list(all_results.keys())

    # Header row: one pair of columns per AI type.
    ws.cell(row=1, column=1, value="Player Strategy")
    for col_idx, ai_label in enumerate(ai_labels):
        ws.cell(row=1, column=2 + col_idx * 2, value=f"{ai_label} Robot WR")
        ws.cell(row=1, column=3 + col_idx * 2, value=f"{ai_label} Streak")
    _format_header_row(ws, 1)

    # Build a lookup dict so we can find combo data by (strategy, ai_label) quickly.
    lookup = {}
    for ai_label, results in all_results.items():
        for combo in results["combo_results"]:
            lookup[(combo["strategy"], ai_label)] = combo

    row = 2
    for strategy in PLAYER_STRATEGIES:
        ws.cell(row=row, column=1, value=strategy).font = BODY_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER

        for col_idx, ai_label in enumerate(ai_labels):
            combo = lookup.get((strategy, ai_label))
            if combo:
                _format_data_cell(ws, row, 2 + col_idx * 2, combo["robot_win_rate"], "pct")
                _format_data_cell(ws, row, 3 + col_idx * 2, combo["avg_streak"], "dec1")
            else:
                # Data missing for this combination — show N/A instead of crashing.
                _format_data_cell(ws, row, 2 + col_idx * 2, "N/A")
                _format_data_cell(ws, row, 3 + col_idx * 2, "N/A")

        row += 1

    _auto_width(ws)


def _build_per_strategy_sheet(wb, all_results):
    """
    Populate 'Per Strategy Detail': one flat row for every strategy+AI combination.

    This gives the full un-averaged data, useful for checking individual combos.
    Robot win rates above 40% are highlighted green; below 28% are red.
    """
    ws = wb.create_sheet("Per Strategy Detail")

    ws.append([
        "Player Strategy",
        "AI Type",
        "Player Win Rate",
        "Robot Win Rate",
        "Draw Rate",
        "Avg Max Streak",
        "Runs",
    ])
    _format_header_row(ws, 1)

    row = 2
    for ai_label, results in all_results.items():
        for combo in results["combo_results"]:
            _format_data_cell(ws, row, 1, combo["strategy"])
            _format_data_cell(ws, row, 2, ai_label)
            _format_data_cell(ws, row, 3, combo["player_win_rate"], "pct")
            _format_data_cell(ws, row, 4, combo["robot_win_rate"],  "pct")
            _format_data_cell(ws, row, 5, combo["draw_rate"],       "pct")
            _format_data_cell(ws, row, 6, combo["avg_streak"], "dec1")
            _format_data_cell(ws, row, 7, combo["runs"])

            # Highlight the robot win rate cell.
            cell = ws.cell(row=row, column=4)
            if combo["robot_win_rate"] > 0.40:
                cell.fill = GOOD_FILL
            elif combo["robot_win_rate"] < 0.28:
                cell.fill = BAD_FILL

            row += 1

    _auto_width(ws)


def _build_key_findings_sheet(wb, all_results, ml_available):
    """
    Populate 'Key Findings' with auto-generated natural-language observations.

    Summarises:
      - Which AI performed best/worst overall
      - How much better each AI was compared to the Random baseline
      - ML vs heuristic comparison (if ML was run)
      - Which player strategy was most exploitable and by which AI
    """
    ws = wb.create_sheet("Key Findings")

    ws.merge_cells("A1:D1")
    ws["A1"]      = "Key Research Findings"
    ws["A1"].font = TITLE_FONT

    findings = []

    # Compute average robot win rate for each AI across all strategies.
    ai_avg = {}
    for ai_label, results in all_results.items():
        combos = results["combo_results"]
        if combos:
            ai_avg[ai_label] = sum(c["robot_win_rate"] for c in combos) / len(combos)

    if ai_avg:
        best_ai  = max(ai_avg, key=ai_avg.get)
        worst_ai = min(ai_avg, key=ai_avg.get)

        findings.append(f"Strongest AI overall: {best_ai} ({ai_avg[best_ai]:.1%} robot win rate)")
        findings.append(f"Weakest AI overall: {worst_ai} ({ai_avg[worst_ai]:.1%} robot win rate)")

        # Show how much each non-random AI beats the random baseline.
        random_wr = ai_avg.get("Random", 0.333)
        for label, wr in ai_avg.items():
            if label != "Random":
                lift = wr - random_wr
                findings.append(f"  {label} lift over random: {lift:+.1%}")

    findings.append("")  # blank line between sections

    # If the ML model was run, compare it directly to each heuristic.
    if ml_available and "ML Prediction" in ai_avg:
        ml_wr = ai_avg["ML Prediction"]
        for label in ["Heuristic (FairPlay)", "Heuristic (Challenge)"]:
            if label in ai_avg:
                h_wr      = ai_avg[label]
                diff      = ml_wr - h_wr
                direction = "outperforms" if diff > 0 else "underperforms"
                findings.append(f"ML {direction} {label} by {abs(diff):.1%}")

    findings.append("")

    # For each strategy, find which AI most exploits it.
    findings.append("Strategy vulnerability analysis:")
    for strategy in PLAYER_STRATEGIES:
        rates = {}
        for ai_label, results in all_results.items():
            for combo in results["combo_results"]:
                if combo["strategy"] == strategy:
                    rates[ai_label] = combo["robot_win_rate"]

        if rates:
            best_against = max(rates, key=rates.get)
            findings.append(
                f"  {strategy}: most exploited by {best_against} ({rates[best_against]:.1%})"
            )

    # Write each finding string to its own row in column A.
    row = 3
    for finding in findings:
        ws.cell(row=row, column=1, value=finding).font = BODY_FONT
        row += 1

    # Wide column so long sentences don't get cut off.
    ws.column_dimensions["A"].width = 70


def _build_ml_details_sheet(wb):
    """
    Populate 'ML Model Details' with metadata and feature importances from the
    trained model file.

    This sheet is only added when ml_available is True (i.e. the .pkl file
    exists).  It gracefully catches any load error and writes it to the cell
    instead of crashing the whole report.
    """
    ws = wb.create_sheet("ML Model Details")

    ws.merge_cells("A1:C1")
    ws["A1"]      = "ML Model Information"
    ws["A1"].font = TITLE_FONT

    try:
        from ml_model import RPSModel
        model = RPSModel.load(ML_MODEL_PATH)

        ws["A3"] = "Model type:"
        ws["B3"] = type(model.model).__name__ if model.model else "N/A"
        ws["A4"] = "Trained:"
        ws["B4"] = "Yes" if model.is_trained else "No"
        ws["A5"] = "Lookback:"
        ws["B5"] = model.lookback

        # Feature importances — useful for understanding what signals the model uses.
        importance = model.get_feature_importance()
        if importance:
            ws["A7"]      = "Feature Importance Ranking"
            ws["A7"].font = SUBTITLE_FONT

            ws["A8"] = "Rank"
            ws["B8"] = "Feature"
            ws["C8"] = "Importance"
            _format_header_row(ws, 8)

            # Write one row per feature, ranked from most to least important.
            for i, (name, score) in enumerate(importance):
                feat_row = 9 + i
                _format_data_cell(ws, feat_row, 1, i + 1)
                _format_data_cell(ws, feat_row, 2, name)
                _format_data_cell(ws, feat_row, 3, round(score, 4))

    except Exception as exc:
        # If anything goes wrong loading the model, record the error in the sheet
        # so the researcher knows what happened without a full crash.
        ws["A3"] = f"Could not load model details: {exc}"

    _auto_width(ws)


# ===========================================================================
# Main entry point
# ===========================================================================

def main():
    """Run all simulations and generate the comparison report."""
    print("=" * 60)
    print("RPS Research Comparison Report Generator")
    print("=" * 60)
    print()

    start = time.time()
    all_results, ml_available = run_all_comparisons()
    elapsed = time.time() - start

    print(f"\nAll simulations complete in {elapsed:.1f}s")

    # Print a quick terminal summary before saving the Excel file.
    print("\n--- AI Performance Summary ---")
    for ai_label, results in all_results.items():
        combos = results["combo_results"]
        if combos:
            avg_rwr = sum(c["robot_win_rate"] for c in combos) / len(combos)
            print(f"  {ai_label:30s}  Robot WR: {avg_rwr:.1%}")

    print()
    build_report(all_results, ml_available, OUTPUT_PATH)


if __name__ == "__main__":
    main()
